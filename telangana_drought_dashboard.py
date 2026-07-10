"""
Telangana AI-Based Drought Prediction and Risk Assessment System
Dashboard v5.0 — District Model Edition
Run:  streamlit run telangana_drought_dashboard.py
Deps: pip install streamlit pandas numpy plotly openpyxl reportlab shap scikit-learn xgboost joblib

Required data files (same folder as this script):
  - District_Drought_Predictions.csv     (output of train_drought_model.py)
  - Telangana_District_Drought_Model.pkl (output of train_drought_model.py)
  - District_Model_Features.pkl          (output of train_drought_model.py)
  - Groundwater_Model.pkl                (optional, separate auxiliary model)
"""

import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from io import BytesIO
import os, datetime, warnings
import re
import difflib
import joblib
import json
import folium
from folium.plugins import Fullscreen, MiniMap, MeasureControl, LocateControl, Search
from streamlit_folium import st_folium
import branca.colormap as bcm

# Optional live GEE fetch — dashboard still works fully without it (falls back
# to manual entry) if the module or its dependencies aren't installed/configured.
try:
    import gee_live_fetch
    GEE_LIVE_AVAILABLE = True
except Exception:
    GEE_LIVE_AVAILABLE = False

warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────────────────────────────────────
#  PAGE CONFIG
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Telangana Drought Prediction System",
    page_icon="🌾", layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────────────────────────────────────
#  CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────
MONTH_ORDER = ['JAN','FEB','MAR','APR','MAY','JUN',
               'JUL','AUG','SEP','OCT','NOV','DEC']
MONTH_FULL  = {'JAN':'January','FEB':'February','MAR':'March','APR':'April',
               'MAY':'May','JUN':'June','JUL':'July','AUG':'August',
               'SEP':'September','OCT':'October','NOV':'November','DEC':'December'}
MONTH_NUM_TO_ABBR = {i+1: m for i, m in enumerate(MONTH_ORDER)}

RISK_ORDER  = ['Low','Moderate','High']
RISK_COLORS = {'High':'#E53E3E','Moderate':'#DD6B20','Low':'#38A169'}
# The district classifier predicts Low / Moderate / High directly (same label set as RiskLevel)
CLASS_COLORS = RISK_COLORS

# Friendlier display names for the "Predicted Class" donut, so it reads as a distinct
# classification (not a visual duplicate of the Risk Level chart below it).
# NOTE: the underlying district model (trained on SPI-3 derived Drought_Label 0/1/2)
# only outputs three drought-severity classes — there is no separate "Wet" class in
# this model's output (that existed only in the old state-level 3-class model).
DISPLAY_CLASS_LABELS = {'Low': 'Normal', 'Moderate': 'Moderate Drought', 'High': 'Severe Drought'}
DISPLAY_CLASS_COLORS = {'Normal': '#38A169', 'Moderate Drought': '#DD6B20', 'Severe Drought': '#E53E3E'}

# Feature labels use the EXACT column names the district model was trained on
# (these must match District_Model_Features.pkl / train_drought_model.py FEATURES list)
FEATURE_LABELS = {
    'Rainfall':            'Rainfall (mm)',
    'Temperature':         'Temperature (°C)',
    'Soil_Moisture':       'Soil Moisture',
    'NDVI':                'NDVI Index',
    'Rainfall_lag1':       'Rainfall — Previous Month (mm)',
    'SoilMoisture_lag1':   'Soil Moisture — Previous Month',
    'NDVI_lag1':           'NDVI — Previous Month',
    'Groundwater_Proxy':   'Groundwater Proxy',
    'SPI3':                'SPI-3 Index',
}

# PLOTLY_BASE without margin — each chart sets its own
PLOTLY_BASE = dict(
    font=dict(family="Segoe UI, sans-serif", size=12, color="#1A202C"),
    plot_bgcolor="#FFFFFF",
    paper_bgcolor="rgba(0,0,0,0)",
    hoverlabel=dict(bgcolor="white", font_size=12, bordercolor="#E2E8F0"),
    xaxis=dict(
        linecolor="#1A202C", linewidth=1.5, mirror=True,
        gridcolor="#E2E8F0", tickfont=dict(color="#1A202C"),
        title_font=dict(color="#1A202C"),
    ),
    yaxis=dict(
        linecolor="#1A202C", linewidth=1.5, mirror=True,
        gridcolor="#E2E8F0", tickfont=dict(color="#1A202C"),
        title_font=dict(color="#1A202C"),
    ),
)

USERS = {
    "admin":   {"password":"drought@2024","role":"Admin",   "name":"Administrator"},
    "officer": {"password":"telangana123","role":"Officer", "name":"District Officer"},
    "viewer":  {"password":"view@2024",   "role":"Viewer",  "name":"Guest Viewer"},
}

DISTRICTS = {
    'Adilabad':(19.664,78.532),

    'Komaram Bheem':(19.333,79.233),      # <-- Change from Kumuram Bheem

    'Mancherial':(18.870,79.459),
    'Nirmal':(19.097,78.343),
    'Nizamabad':(18.672,78.094),
    'Kamareddy':(18.320,78.337),
    'Jagtial':(18.795,78.912),

    'Rajanna Sircilla':(18.386,78.834),   # <-- Fix spelling (Ranjanna -> Rajanna)

    'Karimnagar':(18.438,79.128),
    'Peddapalli':(18.615,79.374),
    'Jayashankar Bhupalapally':(17.865,79.898),
    'Mulugu':(18.193,80.056),
    'Bhadradri Kothagudem':(17.590,80.780),
    'Khammam':(17.247,80.151),
    'Nalgonda':(17.057,79.267),
    'Suryapet':(17.141,79.621),
    'Yadadri Bhuvanagiri':(17.269,78.970),

    'Medchal Malkajgiri':(17.626,78.480),   # <-- Remove hyphen

    'Hyderabad':(17.385,78.486),

    'Ranga Reddy':(17.240,78.400),          # <-- Add space

    'Vikarabad':(17.338,77.902),
    'Sangareddy':(17.627,77.986),
    'Medak':(18.049,78.262),
    'Siddipet':(18.101,78.852),
    'Jangaon':(17.726,79.152),

    'Warangal Rural':(17.978,79.594),       # <-- Remove brackets

    'Warangal Urban':(18.013,79.567),       # <-- Remove brackets

    'Mahabubabad':(17.601,80.004),
    'Nagarkurnool':(16.481,78.324),
    'Wanaparthy':(16.365,78.059),
    'Jogulamba Gadwal':(16.268,78.076),
    'Mahabubnagar':(16.748,77.984),
    'Narayanpet':(16.743,77.494),
}
DISTRICT_LOOKUP_UPPER = {k.upper(): k for k in DISTRICTS}

# ─────────────────────────────────────────────────────────────────────────────
#  GIS MAP CONFIG
# ─────────────────────────────────────────────────────────────────────────────
DISTRICT_GEOJSON_PATH = r"D:\Drought_Temp\telangana_districts.geojson"

# Property keys to try, in order, when reading the district name out of the GeoJSON
GEOJSON_NAME_KEYS = ["District", "district", "DISTRICT", "dtname", "NAME_2", "District_N"]

# Map RAW (UPPERCASE) names found in the GeoJSON to the canonical spelling used in
# DISTRICTS / the CSV. Extend this if your GeoJSON source spells a district differently.
GEOJSON_NAME_MAP = {k.upper(): k for k in DISTRICTS}
GEOJSON_NAME_MAP.update({

    # ---------------------------------------------------------
    # Kumuram Bheem Asifabad
    # ---------------------------------------------------------
    "KOMARAM BHEEM": "Kumuram Bheem",
    "KOMARAM BHEEM ASIFABAD": "Kumuram Bheem",
    "KUMURAM BHEEM": "Kumuram Bheem",
    "KUMURAM BHEEM ASIFABAD": "Kumuram Bheem",
    "KUMRAM BHEEM ASIFABAD": "Kumuram Bheem",
    "K B ASIFABAD": "Kumuram Bheem",
    "ASIFABAD": "Kumuram Bheem",

    # ---------------------------------------------------------
    # Jayashankar Bhupalapally
    # ---------------------------------------------------------
    "JAYASHANKAR": "Jayashankar Bhupalapally",
    "JAYASHANKAR BHUPALAPALLY": "Jayashankar Bhupalapally",
    "JAYASHANKAR BHUPALPALLY": "Jayashankar Bhupalapally",
    "BHUPALAPALLY": "Jayashankar Bhupalapally",
    "BHUPALPALLY": "Jayashankar Bhupalapally",

    # ---------------------------------------------------------
    # Yadadri Bhuvanagiri
    # ---------------------------------------------------------
    "YADADRI": "Yadadri Bhuvanagiri",
    "YADADRI BHUVANAGIRI": "Yadadri Bhuvanagiri",
    "BHUVANAGIRI": "Yadadri Bhuvanagiri",

    # ---------------------------------------------------------
    # Ranga Reddy
    # ---------------------------------------------------------
    "RANGA REDDY": "RangaReddy",
    "RANGAREDDY": "RangaReddy",
    "RANGAREDDI": "RangaReddy",
    "R R DIST": "RangaReddy",

    # ---------------------------------------------------------
    # Medchal-Malkajgiri
    # ---------------------------------------------------------
    "MEDCHAL MALKAJGIRI": "Medchal-Malkajgiri",
    "MEDCHAL-MALKAJGIRI": "Medchal-Malkajgiri",
    "MEDCHAL": "Medchal-Malkajgiri",
    "MALKAJGIRI": "Medchal-Malkajgiri",

    # ---------------------------------------------------------
    # Jogulamba Gadwal
    # ---------------------------------------------------------
    "JOGULAMBA": "Jogulamba Gadwal",
    "JOGULAMBA GADWAL": "Jogulamba Gadwal",
    "GADWAL": "Jogulamba Gadwal",

    # ---------------------------------------------------------
    # Bhadradri Kothagudem
    # ---------------------------------------------------------
    "KOTHAGUDEM": "Bhadradri Kothagudem",
    "BHADRADRI": "Bhadradri Kothagudem",
    "BHADRADRI KOTHAGUDEM": "Bhadradri Kothagudem",

    # ---------------------------------------------------------
    # Rajanna Sircilla
    # ---------------------------------------------------------
    "RAJANNA": "Rajanna Sircilla",
    "RAJANNA SIRCILLA": "Rajanna Sircilla",
    "SIRCILLA": "Rajanna Sircilla",
    "SIRICILLA": "Rajanna Sircilla",

    # ---------------------------------------------------------
    # Warangal
    # ---------------------------------------------------------
    "WARANGAL RURAL": "Warangal (Rural)",
    "WARANGAL (RURAL)": "Warangal (Rural)",
    "WARANGAL URBAN": "Warangal (Urban)",
    "WARANGAL (URBAN)": "Warangal (Urban)",
    "HANAMKONDA": "Warangal (Urban)",
    "HANUMAKONDA": "Warangal (Urban)",

    # ---------------------------------------------------------
    # Mahabubnagar
    # ---------------------------------------------------------
    "MAHABUB NAGAR": "Mahabubnagar",
    "MAHBUBNAGAR": "Mahabubnagar",
    "MAHABOOBNAGAR": "Mahabubnagar",

    # ---------------------------------------------------------
    # Nagarkurnool
    # ---------------------------------------------------------
    "NAGAR KURNOOL": "Nagarkurnool",
    "NAGARKURNOOL": "Nagarkurnool",

    # ---------------------------------------------------------
    # Other districts
    # ---------------------------------------------------------
    "MAHABUBABAD": "Mahabubabad",
    "MAHBUBABAD": "Mahabubabad",

    "JAGITIAL": "Jagtial",
    "JAGTIAL": "Jagtial",

    "PEDDAPALLE": "Peddapalli",
    "PEDDAPALLI": "Peddapalli",

    "NIZAMABAD": "Nizamabad",
    "KAMAREDDY": "Kamareddy",
    "KARIMNAGAR": "Karimnagar",
    "SANGAREDDY": "Sangareddy",
    "SIDDIPET": "Siddipet",
    "VIKARABAD": "Vikarabad",
    "NALGONDA": "Nalgonda",
    "SURYAPET": "Suryapet",
    "KHAMMAM": "Khammam",
    "MULUGU": "Mulugu",
    "NARAYANPET": "Narayanpet",
    "WANAPARTHY": "Wanaparthy",
    "MANCHERIAL": "Mancherial",
    "NIRMAL": "Nirmal",
    "ADILABAD": "Adilabad",
    "MEDAK": "Medak",
    "JANGAON": "Jangaon",
    "HYDERABAD": "Hyderabad",

})

# Basemap tile sources (all free, no API key / no Mapbox)
BASEMAP_TILES = {
    "Street": dict(
        tiles="OpenStreetMap",
        attr="&copy; OpenStreetMap contributors"),
    "Satellite": dict(
        tiles="https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}",
        attr="Tiles &copy; Esri &mdash; Esri, Maxar, Earthstar Geographics"),
    "Terrain": dict(
        tiles="https://{s}.tile.opentopomap.org/{z}/{x}/{y}.png",
        attr="Map data: &copy; OpenStreetMap contributors, SRTM | Map style: &copy; OpenTopoMap (CC-BY-SA)"),
    "Dark": dict(
        tiles="https://{s}.basemaps.cartocdn.com/dark_all/{z}/{x}/{y}{r}.png",
        attr="&copy; OpenStreetMap contributors &copy; CARTO"),
    "Light": dict(
        tiles="https://{s}.basemaps.cartocdn.com/light_all/{z}/{x}/{y}{r}.png",
        attr="&copy; OpenStreetMap contributors &copy; CARTO"),
}

# Continuous indicators that can drive polygon fill color, besides risk level
MAP_INDICATOR_OPTIONS = {
    "Prediction Risk":     ("RiskLevel", "discrete"),
    "Rainfall":            ("Rainfall", "continuous"),
    "Temperature":         ("Temperature", "continuous"),
    "NDVI":                ("NDVI", "continuous"),
    "SPI-3":               ("SPI3", "continuous"),
    "Groundwater Proxy":   ("Groundwater_Proxy", "continuous"),
}

# ─────────────────────────────────────────────────────────────────────────────
#  CSS
# ─────────────────────────────────────────────────────────────────────────────
def inject_css():
    st.markdown("""
<style>
/* ── Reset & globals ── */
html,body,[class*="css"]{ font-family:'Segoe UI',Arial,sans-serif; }

/* ── App background ── */
.stApp {
    background:
        radial-gradient(circle at 8% 8%,  rgba(43,108,176,0.07), transparent 42%),
        radial-gradient(circle at 92% 92%,rgba(44,122,123,0.07), transparent 42%),
        #E8EDF2 !important;
    min-height:100vh;
}

/* ── Main content panel ── */
.main .block-container {
    background: transparent !important;
    padding: 16px 24px 40px !important;
    max-width: 1400px;
}

/* ── Section cards — solid white, dark text ── */
.section-card {
    background: #FFFFFF !important;
    border-radius: 12px;
    padding: 20px 22px;
    margin-bottom: 18px;
    box-shadow: 0 4px 24px rgba(0,0,0,0.10);
    border: 1px solid #E2E8F0;
}
.section-card p, .section-card span,
.section-card div, .section-card label { color: #1A202C !important; }
.section-card iframe { border-radius: 14px !important; }
div[data-testid="stVerticalBlock"] .folium-map,
div[data-testid="stVerticalBlock"] iframe[title="streamlit_folium.st_folium"] {
    border-radius: 14px !important;
    box-shadow: 0 6px 20px rgba(26,54,93,0.12) !important;
    border: 1px solid #E2E8F0 !important;
}
.section-title {
    font-family:'Segoe UI',Georgia,serif !important;
    font-size:1.0rem !important; font-weight:700 !important;
    color:#1A365D !important; margin:0 0 12px !important;
    padding-bottom:10px !important;
    border-bottom:2px solid #BEE3F8 !important;
    background: none !important;
    -webkit-text-fill-color: #1A365D !important;
}

/* ── KPI cards ── */
.kpi-card { background: #FFFFFF !important; }
.kpi-label { color: #4A5568 !important; }
.kpi-value { color: #1A365D !important; }
.kpi-sub   { color: #718096 !important; }

/* ═══════════════════════════════════
   LOGIN PAGE
═══════════════════════════════════ */
.login-hero {
    text-align:center; padding: 26px 10px 6px;
}
.login-hero .lh-eyebrow {
    display:inline-block; background:rgba(43,108,176,0.10);
    color:#2B6CB0; font-size:0.68rem; font-weight:700;
    letter-spacing:1.4px; text-transform:uppercase;
    padding:5px 16px; border-radius:20px; margin-bottom:14px;
}
.login-hero h2 {
    color:#1A365D; font-family:'Segoe UI',Georgia,serif;
    font-size:1.9rem; font-weight:800; margin:0 0 8px;
}
.login-hero p {
    color:#4A5568; font-size:0.88rem; max-width:560px; margin:0 auto;
}

.login-stat-chip {
    background:rgba(255,255,255,0.85); backdrop-filter:blur(6px);
    border:1px solid rgba(99,179,237,0.30); border-radius:14px;
    padding:16px 14px; text-align:center; margin: 10px 4px 22px;
    box-shadow:0 6px 20px rgba(26,54,93,0.08);
    transition:transform .2s;
}
.login-stat-chip:hover{ transform:translateY(-3px); }
.lsc-icon { font-size:1.5rem; display:block; margin-bottom:4px; }
.lsc-val  { font-size:1.05rem; font-weight:800; color:#1A365D; }
.lsc-label{ font-size:0.66rem; color:#718096; font-weight:600;
            letter-spacing:0.4px; text-transform:uppercase; margin-top:2px; }

.login-bg {
    min-height: 90vh;
    display: flex; align-items: center; justify-content: center;
}
.login-outer { display:flex; align-items:center; justify-content:center; min-height:60vh; }
.login-card {
    background: linear-gradient(160deg, rgba(255,255,255,0.98) 0%, rgba(235,248,255,0.98) 100%);
    border-radius: 28px;
    padding: 50px 54px 44px;
    box-shadow:
        0 40px 100px rgba(0,0,0,0.30),
        0 0 0 1px rgba(99,179,237,0.15),
        inset 0 1px 0 rgba(255,255,255,0.8);
    max-width: 480px; width: 100%;
    text-align: center;
    position: relative; overflow: hidden;
}
/* Animated top rainbow bar */
.login-card::before {
    content:'';
    position:absolute; top:0; left:0; right:0; height:6px;
    background: linear-gradient(90deg,#1A365D,#2B6CB0,#38B2AC,#3182CE,#2C7A7B,#1A365D);
    background-size:300% 100%;
    animation: shimmer 4s linear infinite;
}
/* Subtle watermark circle */
.login-card::after {
    content:'🌾';
    position:absolute; bottom:-30px; right:-20px;
    font-size:9rem; opacity:0.04; transform:rotate(-15deg);
    pointer-events:none;
}
@keyframes shimmer { 0%{background-position:300% 0} 100%{background-position:-300% 0} }
.login-emblem {
    font-size: 4rem; display:block; margin-bottom:8px;
    filter: drop-shadow(0 6px 12px rgba(43,108,176,0.35));
    animation: float 3s ease-in-out infinite;
}
@keyframes float {
    0%,100%{transform:translateY(0)} 50%{transform:translateY(-6px)}
}
.login-card h1 {
    color: #1A365D; font-family:'Segoe UI',Georgia,serif;
    font-size: 1.65rem; font-weight:800; margin:10px 0 4px; line-height:1.2;
    background:linear-gradient(135deg,#1A365D,#2B6CB0);
    -webkit-background-clip:text; -webkit-text-fill-color:transparent; background-clip:text;
}
.login-card .login-dept {
    color:#4A5568; font-size:0.80rem; margin:0 0 28px; line-height:1.6;
}
.login-card .login-badge {
    display:inline-block;
    background:linear-gradient(90deg,#1A365D,#2B6CB0);
    color:#fff; font-size:0.63rem; font-weight:700; letter-spacing:1.2px;
    text-transform:uppercase; padding:4px 14px; border-radius:20px;
    margin-bottom:12px; box-shadow:0 4px 12px rgba(43,108,176,0.30);
}
/* Login form divider line */
.login-divider {
    display:flex; align-items:center; gap:10px;
    margin:16px 0 8px; color:#A0AEC0; font-size:0.73rem;
}
.login-divider::before,.login-divider::after {
    content:''; flex:1; height:1px; background:rgba(0,0,0,0.10);
}
.login-forgot { text-align:right; margin-top:6px; }
.login-forgot a { color:#3182CE; font-size:0.72rem; text-decoration:none; font-weight:600; }
.login-forgot a:hover { text-decoration:underline; }

/* ═══════════════════════════════════
   SIDEBAR
═══════════════════════════════════ */
section[data-testid="stSidebar"] {
    background: linear-gradient(180deg,#071428 0%,#0d1f3c 30%,#0f2744 60%,#0d2137 100%) !important;
    border-right: 1px solid rgba(99,179,237,0.20) !important;
}
section[data-testid="stSidebar"] > div { padding-top:0 !important; }

.sb-header {
    background: linear-gradient(135deg,rgba(49,130,206,0.25),rgba(43,108,176,0.15));
    border-bottom: 1px solid rgba(99,179,237,0.15);
    padding: 20px 16px 16px; text-align:center;
}
.sb-icon { font-size:2.8rem; display:block; margin-bottom:6px; }
.sb-title {
    color:#EBF8FF !important; font-family:'Segoe UI',Georgia,serif !important;
    font-size:0.90rem !important; font-weight:700 !important; line-height:1.3 !important;
    margin:0 0 3px !important;
}
.sb-sub { color:rgba(190,227,248,0.55) !important; font-size:0.63rem !important; margin:0 !important; }

.sb-user {
    background:rgba(99,179,237,0.08);
    border:1px solid rgba(99,179,237,0.18);
    border-radius:10px; padding:10px 12px; margin:12px 8px 4px;
    display:flex; align-items:center; gap:10px;
}
.sb-av {
    background:linear-gradient(135deg,#3182CE,#63B3ED);
    color:#fff; border-radius:50%; width:32px; height:32px;
    display:flex; align-items:center; justify-content:center;
    font-weight:700; font-size:0.82rem; flex-shrink:0;
    box-shadow:0 2px 8px rgba(49,130,206,0.4);
}
.sb-uname { color:#EBF8FF !important; font-size:0.80rem !important; font-weight:600 !important; line-height:1.2 !important; }
.sb-urole { color:rgba(190,227,248,0.50) !important; font-size:0.65rem !important; }

.sb-sec {
    background: rgba(99,179,237,0.08);
    border-left:3px solid #3182CE;
    border-radius:0 6px 6px 0;
    padding:5px 12px; margin:14px 0 6px;
    font-size:0.67rem !important; font-weight:700 !important;
    letter-spacing:1.2px !important; text-transform:uppercase !important;
    color:#63B3ED !important;
}
.sb-divider { border:none; border-top:1px solid rgba(99,179,237,0.12); margin:10px 0; }

/* Fix multiselect tags visible on dark sidebar */
section[data-testid="stSidebar"] * { color:rgba(235,248,255,0.88) !important; }
section[data-testid="stSidebar"] label {
    font-size:0.73rem !important; font-weight:600 !important;
    color:rgba(190,227,248,0.75) !important; letter-spacing:0.2px !important;
}
/* Fix logout button — white text on dark button */
section[data-testid="stSidebar"] .stButton > button {
    background: rgba(229,62,62,0.18) !important;
    border: 1px solid rgba(229,62,62,0.40) !important;
    color: #FEB2B2 !important;
    font-weight: 600 !important;
    border-radius: 8px !important;
}
section[data-testid="stSidebar"] .stButton > button:hover {
    background: rgba(229,62,62,0.32) !important;
    color: #fff !important;
}
section[data-testid="stSidebar"] .stMultiSelect [data-baseweb="tag"] {
    background:rgba(49,130,206,0.85) !important;
    border-color:rgba(49,130,206,0.85) !important;
}
section[data-testid="stSidebar"] .stMultiSelect [data-baseweb="tag"] span,
section[data-testid="stSidebar"] .stMultiSelect [data-baseweb="tag"] button {
    color:#ffffff !important;
}
section[data-testid="stSidebar"] .stMultiSelect [data-baseweb="select"]>div,
section[data-testid="stSidebar"] .stSelectbox [data-baseweb="select"]>div {
    background:rgba(15,39,68,0.92) !important;
    border-color:rgba(99,179,237,0.30) !important;
}
/* Force readable light text on the dark selectbox background — the selected value
   display was previously rendering white-on-white/near-invisible because a default
   BaseWeb background was winning over our translucent override. */
section[data-testid="stSidebar"] .stSelectbox [data-baseweb="select"],
section[data-testid="stSidebar"] .stSelectbox [data-baseweb="select"] * {
    background-color: transparent !important;
    color: #EBF8FF !important;
}
section[data-testid="stSidebar"] .stSelectbox [data-baseweb="select"] > div {
    background-color: rgba(15,39,68,0.92) !important;
}
section[data-testid="stSidebar"] .stSelectbox svg {
    fill: #EBF8FF !important;
}

/* ═══════════════════════════════════
   DASHBOARD ELEMENTS
═══════════════════════════════════ */

/* Header banner */
.gov-header {
    background: linear-gradient(135deg,#1A365D 0%,#2B6CB0 50%,#2C7A7B 100%);
    padding:22px 30px 18px; border-radius:14px; margin-bottom:20px;
    position:relative; overflow:hidden;
    box-shadow:0 8px 32px rgba(43,108,176,0.35);
    border:1px solid rgba(99,179,237,0.20);
}
.gov-header::before {
    content:''; position:absolute; top:-50%; right:-10%;
    width:400px; height:400px; border-radius:50%;
    background:radial-gradient(circle,rgba(99,179,237,0.12) 0%,transparent 70%);
}
.gov-header::after {
    content:'🌾'; position:absolute; right:28px; top:50%;
    transform:translateY(-50%); font-size:4rem; opacity:0.08;
}
.gov-badge {
    display:inline-block;
    background:rgba(255,255,255,0.15); backdrop-filter:blur(4px);
    border:1px solid rgba(255,255,255,0.25);
    color:#BEE3F8; font-size:0.64rem; font-weight:600;
    letter-spacing:0.8px; text-transform:uppercase;
    padding:3px 12px; border-radius:20px; margin-bottom:8px;
}
.gov-header h1 {
    color:#fff; font-family:'Segoe UI',Georgia,serif;
    font-size:1.60rem; font-weight:800; margin:0 0 5px; letter-spacing:-0.3px;
    text-shadow:0 2px 8px rgba(0,0,0,0.2);
}
.gov-header p { color:rgba(255,255,255,0.72); font-size:0.80rem; margin:0; line-height:1.55; }

/* KPI cards */
.kpi-card {
    background:rgba(255,255,255,0.97);
    border-radius:12px; padding:18px 20px;
    box-shadow:0 4px 24px rgba(0,0,0,0.12);
    border-top:4px solid #2B6CB0;
    transition:transform 0.2s,box-shadow 0.2s;
    position:relative; overflow:hidden;
}
.kpi-card::after {
    content:''; position:absolute; bottom:-20px; right:-20px;
    width:80px; height:80px; border-radius:50%;
    background:rgba(43,108,176,0.05);
}
.kpi-card:hover { transform:translateY(-4px); box-shadow:0 12px 32px rgba(0,0,0,0.18); }
.kpi-card.red   { border-top-color:#E53E3E; }
.kpi-card.blue  { border-top-color:#3182CE; }
.kpi-card.teal  { border-top-color:#2C7A7B; }
.kpi-card.orange{ border-top-color:#DD6B20; }
.kpi-card.green { border-top-color:#38A169; }
.kpi-card.rust  { border-top-color:#B84A20; }
.kpi-card.amber { border-top-color:#C47B1A; }
.kpi-label { font-size:0.68rem; font-weight:600; letter-spacing:0.7px;
             text-transform:uppercase; color:#718096; margin-bottom:7px; }
.kpi-value { font-size:2.0rem; font-weight:800; color:#1A365D; line-height:1; }
.kpi-sub   { font-size:0.72rem; color:#A0AEC0; margin-top:5px; }

/* Tabs — sit on white/light background so text must be dark */
.stTabs [data-baseweb="tab-list"] {
    background:#EBF8FF; border-radius:10px;
    padding:3px; gap:2px;
    border:1px solid #BEE3F8;
}
.stTabs [data-baseweb="tab"] {
    border-radius:8px; padding:8px 18px;
    font-size:0.80rem; font-weight:600; color:#2D3748;
}
.stTabs [aria-selected="true"] {
    background:linear-gradient(135deg,#2B6CB0,#3182CE) !important;
    color:white !important;
    box-shadow:0 4px 12px rgba(49,130,206,0.35) !important;
}

/* Main navigation (lazy-render radio standing in for tabs) — pill styling */
div[data-testid="stRadio"] > div[role="radiogroup"] {
    background:#EBF8FF; border-radius:10px;
    padding:4px; gap:4px;
    border:1px solid #BEE3F8;
    flex-wrap:wrap;
}
div[data-testid="stRadio"] label {
    border-radius:8px !important; padding:8px 16px !important;
    font-size:0.80rem !important; font-weight:600 !important;
    color:#2D3748 !important; margin:2px !important;
    background:transparent !important;
    transition:background 0.15s, color 0.15s;
}
div[data-testid="stRadio"] label:hover {
    background:rgba(43,108,176,0.10) !important;
}
div[data-testid="stRadio"] label:has(input:checked) {
    background:linear-gradient(135deg,#2B6CB0,#3182CE) !important;
    box-shadow:0 4px 12px rgba(49,130,206,0.35) !important;
}
div[data-testid="stRadio"] label:has(input:checked) p {
    color:#ffffff !important; font-weight:700 !important;
}
div[data-testid="stRadio"] label > div:first-child {
    display:none;   /* hide the default radio circle, pill background shows selection instead */
}

/* Lookup boxes */
.lookup-box {
    background:linear-gradient(135deg,#EBF8FF 0%,#E6FFFA 100%);
    border:1px solid #BEE3F8; border-left:5px solid #3182CE;
    border-radius:10px; padding:18px 22px; margin-top:10px;
    box-shadow:0 2px 12px rgba(49,130,206,0.08);
}
.lookup-box h4 {
    color:#1A365D; font-family:'Segoe UI',Georgia,serif;
    margin:0 0 12px; font-size:1.0rem;
}
.lrow { display:flex; justify-content:space-between; align-items:flex-start;
        padding:7px 0; border-bottom:1px solid rgba(49,130,206,0.10); }
.lrow:last-child { border-bottom:none; }
.lkey { font-size:0.78rem; color:#718096; font-weight:500; min-width:130px; }
.lval { font-size:0.90rem; color:#1A365D; font-weight:600;
        text-align:right; max-width:68%; line-height:1.4; }

/* Alert boxes */
.alert-high { background:linear-gradient(135deg,#FFF5F5,#FED7D7);
              border:1px solid #FEB2B2; border-left:5px solid #E53E3E;
              border-radius:10px; padding:14px 18px; margin:10px 0;
              box-shadow:0 2px 12px rgba(229,62,62,0.10); }
.alert-mod  { background:linear-gradient(135deg,#FFFAF0,#FEEBC8);
              border:1px solid #FBD38D; border-left:5px solid #DD6B20;
              border-radius:10px; padding:14px 18px; margin:10px 0; }
.alert-low  { background:linear-gradient(135deg,#F0FFF4,#C6F6D5);
              border:1px solid #9AE6B4; border-left:5px solid #38A169;
              border-radius:10px; padding:14px 18px; margin:10px 0; }

footer { visibility:hidden; }
#MainMenu { visibility:hidden; }
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
#  LOGIN
# ─────────────────────────────────────────────────────────────────────────────
def show_login():
    inject_css()

    st.markdown("""
    <div class="login-hero">
        <span class="lh-eyebrow">🏛️ Government of Telangana · Dept. of Agriculture &amp; Cooperation</span>
        <h2>Telangana District-Level Drought<br>Prediction &amp; Risk Assessment</h2>
        <p>An AI/ML early-warning platform combining satellite, meteorological and groundwater
           indicators across all 33 districts of Telangana — powered by an ensemble
           Random Forest + XGBoost stacking model with SHAP explainability.</p>
    </div>
    """, unsafe_allow_html=True)

    sc1, sc2, sc3 = st.columns(3)
    stats = [
        ("🗺️", "33", "Districts Monitored"),
        ("🤖", "RF + XGB", "Stacking Ensemble Model"),
        ("🧬", "SHAP", "Explainable AI Insights"),
    ]
    for col, (icon, val, label) in zip([sc1, sc2, sc3], stats):
        with col:
            st.markdown(f"""
            <div class="login-stat-chip">
                <span class="lsc-icon">{icon}</span>
                <div class="lsc-val">{val}</div>
                <div class="lsc-label">{label}</div>
            </div>""", unsafe_allow_html=True)

    _, col_c, _ = st.columns([1, 1.4, 1])
    with col_c:
        st.markdown("""
        <div class="login-outer">
        <div class="login-card">
            <span class="login-emblem">🌾</span>
            <span class="login-badge">🔒 Authorised Personnel Only</span>
            <h1>Sign In to the<br>District Dashboard</h1>
            <p class="login-dept">
                Real-time district drought risk, live predictions and advisories<br>
                <span style="color:#718096;">AI/ML Early Warning System &nbsp;·&nbsp; District Model v5.0</span>
            </p>
            <div class="login-divider">Sign in with your credentials</div>
        </div>
        </div>""", unsafe_allow_html=True)

        with st.form("login_form"):
            username  = st.text_input("👤 Username", placeholder="Enter your username")
            password  = st.text_input("🔑 Password", type="password", placeholder="Enter your password")
            remember  = st.checkbox("Remember me on this device", value=True)
            submitted = st.form_submit_button("🔐  Sign In to Dashboard", use_container_width=True)

        st.markdown('<div class="login-forgot"><a href="#">Forgot password? Contact your administrator</a></div>',
                    unsafe_allow_html=True)

        if submitted:
            if username in USERS and USERS[username]["password"] == password:
                st.session_state.update({
                    "logged_in":  True,
                    "username":   username,
                    "user_name":  USERS[username]["name"],
                    "user_role":  USERS[username]["role"],
                    "login_time": datetime.datetime.now().strftime("%d %b %Y, %H:%M"),
                })
                st.rerun()
            else:
                st.error("❌ Invalid credentials. Please check your username and password.")

        st.markdown("""
        <div style='text-align:center;margin-top:14px;font-size:0.68rem;color:#A0AEC0;
                    padding:10px 14px;background:rgba(0,0,0,0.04);border-radius:10px;
                    border:1px solid rgba(0,0,0,0.06);'>
            Demo credentials &nbsp;·&nbsp;
            <strong>admin</strong> / drought@2024 &nbsp;·&nbsp;
            <strong>officer</strong> / telangana123 &nbsp;·&nbsp;
            <strong>viewer</strong> / view@2024
        </div>""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
#  DATA LOADER — real district-level model predictions
# ─────────────────────────────────────────────────────────────────────────────
DISTRICT_CSV_PATH = r"D:\Drought_Temp\District_Drought_Predictions.csv"

@st.cache_data(show_spinner=False)
def load_district_predictions(path):
    """
    Loads the REAL district-level prediction output produced by
    train_drought_model.py (District_Drought_Predictions.csv).

    Expected columns include (names must match the training script):
      District, Year, Month, Rainfall, Temperature, Soil_Moisture, NDVI,
      Rainfall_lag1, SoilMoisture_lag1, NDVI_lag1, Groundwater_Proxy, SPI3,
      Predicted_Class (Low/Moderate/High), Confidence, RiskScore, RiskLevel, Advisory
    """
    df = pd.read_csv(path)
    df.columns = [c.strip() for c in df.columns]

    # Month may be numeric (1-12) as produced by the training script, or already text
    if pd.api.types.is_numeric_dtype(df['Month']):
        df['Month_Num'] = df['Month'].astype(int)
        df['Month'] = df['Month_Num'].map(MONTH_NUM_TO_ABBR)
    else:
        df['Month'] = df['Month'].astype(str).str.strip().str.upper().str[:3]
        df['Month_Num'] = df['Month'].apply(lambda m: MONTH_ORDER.index(m)+1 if m in MONTH_ORDER else 0)

    # Normalise District names against the canonical DISTRICTS lat/lon table
    df['District'] = df['District'].astype(str).str.strip()
    df['District'] = df['District'].apply(lambda d: DISTRICT_LOOKUP_UPPER.get(d.upper(), d))
    df['Lat'] = df['District'].map(lambda d: DISTRICTS.get(d, (None, None))[0])
    df['Lon'] = df['District'].map(lambda d: DISTRICTS.get(d, (None, None))[1])

    df = df.sort_values(['District','Year','Month_Num']).reset_index(drop=True)
    df['Period']     = df['Month'] + ' ' + df['Year'].astype(str)
    df['Month_Full'] = df['Month'].map(MONTH_FULL).fillna(df['Month'])

    # Safety: make sure RiskLevel / Predicted_Class exist even if the CSV is missing them
    if 'RiskLevel' not in df.columns and 'RiskScore' in df.columns:
        df['RiskLevel'] = df['RiskScore'].apply(lambda s: 'High' if s>=70 else ('Moderate' if s>=40 else 'Low'))
    if 'Predicted_Class' not in df.columns and 'RiskLevel' in df.columns:
        df['Predicted_Class'] = df['RiskLevel']
    if 'Advisory' not in df.columns:
        df['Advisory'] = ''

    return df


# ─────────────────────────────────────────────────────────────────────────────
#  MODEL LOADERS
# ─────────────────────────────────────────────────────────────────────────────
DISTRICT_MODEL_PATH    = r"D:\Drought_Temp\Telangana_District_Drought_Model.pkl"
DISTRICT_FEATURES_PATH = r"D:\Drought_Temp\District_Model_Features.pkl"
# Fallback list — MUST mirror the FEATURES list/order in train_drought_model.py exactly
FALLBACK_DISTRICT_FEATURES = [
    "Rainfall", "Temperature", "Soil_Moisture", "NDVI",
    "Rainfall_lag1", "SoilMoisture_lag1", "NDVI_lag1", "Groundwater_Proxy",
]
DISTRICT_CLASS_LABEL_MAP = {0: 'Low', 1: 'Moderate', 2: 'High'}

# ── Groundwater Proxy Model (auxiliary RandomForestRegressor) ────────────────
GW_MODEL_PATH     = "Groundwater_Model.pkl"
GW_MODEL_FEATURES = ["Rainfall", "Temperature", "SoilMoisture", "NDVI"]  # this model's own training order

@st.cache_resource(show_spinner=False)
def load_district_model(path: str):
    if not os.path.exists(path):
        return None
    try:
        return joblib.load(path)
    except Exception as e:
        st.error(f"Failed to load district drought model: {e}")
        return None

@st.cache_resource(show_spinner=False)
def load_district_features(path: str):
    if os.path.exists(path):
        try:
            feats = joblib.load(path)
            return list(feats)
        except Exception:
            pass
    return FALLBACK_DISTRICT_FEATURES

@st.cache_resource(show_spinner=False)
def load_gw_model():
    if not os.path.exists(GW_MODEL_PATH):
        return None
    try:
        return joblib.load(GW_MODEL_PATH)
    except Exception as e:
        st.error(f"Failed to load Groundwater model: {e}")
        return None

def _label_for(raw_pred):
    """Map a raw model prediction (int 0/1/2 or already a string) to Low/Moderate/High."""
    if isinstance(raw_pred, (int, np.integer)):
        return DISTRICT_CLASS_LABEL_MAP.get(int(raw_pred), str(raw_pred))
    if isinstance(raw_pred, str) and raw_pred not in RISK_ORDER:
        try:
            return DISTRICT_CLASS_LABEL_MAP.get(int(raw_pred), raw_pred)
        except ValueError:
            return raw_pred
    return raw_pred

def run_prediction(model, feature_row: dict, features_list: list):
    """
    feature_row: dict of {feature_name: value} using the EXACT names in features_list.
    Returns: (predicted_class_label, confidence, risk_score)
    """
    X = pd.DataFrame([feature_row])[features_list]
    raw_pred = model.predict(X)[0]
    pred = _label_for(raw_pred)

    confidence = None
    if hasattr(model, "predict_proba"):
        proba = model.predict_proba(X)[0]
        confidence = float(np.max(proba))

    # Risk score heuristic: scale confidence by class severity (High = worst)
    risk_map = {"High": 1.0, "Moderate": 0.5, "Low": 0.15}
    base_risk = risk_map.get(str(pred), 0.5)
    risk_score = round(base_risk * (confidence if confidence else 0.7) * 100, 1)

    return pred, confidence, risk_score

def run_gw_prediction(gw_model, feature_row: dict):
    """
    Runs the Groundwater_Model.pkl (RandomForestRegressor).
    Input : Rainfall, Temperature, SoilMoisture, NDVI
    Output: Groundwater_Proxy float 0-1
    """
    X = pd.DataFrame([feature_row])[GW_MODEL_FEATURES]
    gw_proxy = float(gw_model.predict(X)[0])
    gw_proxy = round(np.clip(gw_proxy, 0.0, 1.0), 4)

    if gw_proxy >= 0.66:
        gw_status, gw_color = "Adequate", "#38A169"
    elif gw_proxy >= 0.33:
        gw_status, gw_color = "Moderate", "#DD6B20"
    else:
        gw_status, gw_color = "Critically Low", "#E53E3E"

    return gw_proxy, gw_status, gw_color

# ─────────────────────────────────────────────────────────────────────────────
#  SHAP EXPLAINABILITY
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_resource(show_spinner=False)
def build_shap_explainer(_model, _background_df: pd.DataFrame, _features_list: list):
    """
    Model-agnostic SHAP KernelExplainer — safe for stacking ensembles
    (RF + XGBoost -> Logistic Regression), since there's no single tree
    structure a TreeExplainer could use directly across the whole stack.
    """
    import shap
    background = _background_df[_features_list].dropna().sample(
        n=min(40, len(_background_df)), random_state=42
    )
    explainer = shap.KernelExplainer(_model.predict_proba, background)
    return explainer

def explain_prediction(explainer, model, feature_row: dict, features_list: list):
    X = pd.DataFrame([feature_row])[features_list]
    raw_pred = model.predict(X)[0]
    class_list = list(model.classes_)
    class_idx  = class_list.index(raw_pred) if raw_pred in class_list else 0
    pred_label = _label_for(raw_pred)

    shap_values = explainer.shap_values(X, nsamples=100)

    if isinstance(shap_values, list):
        sv = shap_values[class_idx][0]
    else:
        sv = shap_values[0, :, class_idx]

    result = pd.DataFrame({
        'Feature':    [FEATURE_LABELS.get(f, f) for f in features_list],
        'Value':      [feature_row[f] for f in features_list],
        'SHAP_Value': sv,
    })
    result['AbsImpact'] = result['SHAP_Value'].abs()
    result = result.sort_values('AbsImpact', ascending=False).reset_index(drop=True)
    return result, pred_label


# ─────────────────────────────────────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def compute_feature_importance(df):
    feat_cols = [c for c in FEATURE_LABELS if c in df.columns]
    corrs     = {FEATURE_LABELS[c]: abs(df[c].corr(df['RiskScore'])) for c in feat_cols}
    fi        = pd.DataFrame({'Feature':list(corrs.keys()),'Importance':list(corrs.values())})
    fi        = fi.dropna().sort_values('Importance')
    total     = fi['Importance'].sum()
    if total > 0: fi['Importance'] = fi['Importance']/total
    return fi

def to_excel_bytes(df):
    cols = [c for c in ['District','Year','Month','Predicted_Class','Confidence','RiskScore',
                         'RiskLevel','Rainfall','Temperature','Soil_Moisture',
                         'NDVI','SPI3','Groundwater_Proxy','Advisory'] if c in df.columns]
    out = BytesIO()
    with pd.ExcelWriter(out, engine='openpyxl') as writer:
        df[cols].to_excel(writer, index=False, sheet_name='Drought Predictions')
        ws  = writer.sheets['Drought Predictions']
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        hf  = PatternFill('solid', fgColor='1A365D')
        hft = Font(bold=True, color='FFFFFF', size=10)
        thn = Border(left=Side(style='thin',color='E2E8F0'),
                     right=Side(style='thin',color='E2E8F0'),
                     bottom=Side(style='thin',color='E2E8F0'))
        widths = {'District':20,'Year':7,'Month':8,'Predicted_Class':16,'Confidence':13,
                  'RiskScore':11,'RiskLevel':12,'Rainfall':11,'Temperature':13,
                  'Soil_Moisture':13,'NDVI':9,'SPI3':9,'Groundwater_Proxy':16,'Advisory':50}
        for cell in ws[1]:
            cell.fill=hf; cell.font=hft
            cell.alignment=Alignment(horizontal='center',wrap_text=True)
        for i, col in enumerate(cols,1):
            ws.column_dimensions[ws.cell(1,i).column_letter].width=widths.get(col,14)
            for row in ws.iter_rows(min_row=2,min_col=i,max_col=i):
                for cell in row:
                    cell.border=thn
                    cell.alignment=Alignment(horizontal='center')
    return out.getvalue()

def generate_pdf_report(df, sel_years, sel_months, sel_districts, username):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                        Table, TableStyle, HRFlowable)
        from reportlab.lib.enums import TA_CENTER

        buf    = BytesIO()
        doc    = SimpleDocTemplate(buf,pagesize=A4,
                                    rightMargin=1.6*cm,leftMargin=1.6*cm,
                                    topMargin=2*cm,bottomMargin=2*cm)
        styles = getSampleStyleSheet()
        navy   = colors.HexColor('#1A365D')
        blue   = colors.HexColor('#2B6CB0')
        red    = colors.HexColor('#E53E3E')
        orange = colors.HexColor('#DD6B20')
        green  = colors.HexColor('#38A169')

        T = ParagraphStyle('T',parent=styles['Title'],textColor=navy,fontSize=16,
                            spaceAfter=4,alignment=TA_CENTER)
        S = ParagraphStyle('S',parent=styles['Normal'],
                            textColor=colors.HexColor('#718096'),
                            fontSize=9,spaceAfter=14,alignment=TA_CENTER)
        H = ParagraphStyle('H',parent=styles['Heading2'],textColor=navy,
                            fontSize=11,spaceBefore=12,spaceAfter=6)
        B = ParagraphStyle('B',parent=styles['Normal'],fontSize=9,leading=13,
                            textColor=colors.HexColor('#2D3748'))

        story=[]
        story.append(Paragraph("Telangana AI-Based District Drought Prediction &amp; Risk Assessment System",T))
        story.append(Paragraph(
            f"Official Report &nbsp;·&nbsp; "
            f"Generated: {datetime.datetime.now().strftime('%d %b %Y %H:%M')} &nbsp;·&nbsp; By: {username}",S))
        story.append(HRFlowable(width="100%",thickness=2,color=blue,spaceAfter=10))

        story.append(Paragraph("Report Parameters",H))
        yrs  = ', '.join(str(y) for y in sel_years) if sel_years else 'All'
        mnts = ', '.join(MONTH_FULL.get(m,m) for m in sel_months) if sel_months else 'All'
        dstr = ', '.join(sel_districts) if sel_districts else 'All (33 Districts)'
        story.append(Paragraph(f"<b>Years:</b> {yrs} &nbsp; <b>Months:</b> {mnts} &nbsp; <b>Records:</b> {len(df)}",B))
        story.append(Paragraph(f"<b>Districts:</b> {dstr}",B))
        story.append(Spacer(1,8))

        story.append(Paragraph("Summary Statistics",H))
        h_n  = int((df['Predicted_Class']=='High').sum())
        mo_n = int((df['Predicted_Class']=='Moderate').sum())
        l_n  = int((df['Predicted_Class']=='Low').sum())
        ac   = df['Confidence'].mean()*100
        ar   = df['RiskScore'].mean()
        dist_n = df['District'].nunique() if 'District' in df.columns else 0

        kdata=[['Metric','Value','Metric','Value'],
               ['High Risk Records',str(h_n),'Moderate Risk Records',str(mo_n)],
               ['Low Risk Records',str(l_n),'Districts Covered',str(dist_n)],
               ['Avg. Confidence',f'{ac:.1f}%','Avg. Risk Score',f'{ar:.1f}/100']]
        ktbl=Table(kdata,colWidths=[4.4*cm,2.8*cm,4.4*cm,2.8*cm])
        ktbl.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0),navy),('TEXTCOLOR',(0,0),(-1,0),colors.white),
            ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,-1),9),
            ('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('GRID',(0,0),(-1,-1),0.5,colors.HexColor('#E2E8F0')),('ROWHEIGHT',(0,0),(-1,-1),20),
        ]))
        story.append(ktbl)
        story.append(Spacer(1,12))

        story.append(Paragraph("District-Month Prediction Details",H))
        df_sorted=df.sort_values(['District','Year','Month_Num']) if 'Month_Num' in df.columns else df.sort_values('District')
        show=df_sorted[['District','Year','Month','Predicted_Class','Confidence','RiskScore','RiskLevel']].copy()
        show['Confidence']=(show['Confidence']*100).round(1).astype(str)+'%'
        show['RiskScore']=show['RiskScore'].round(1)
        # cap rows for a sane PDF size
        show = show.head(500)

        tdata=[['District','Year','Month','Prediction','Confidence','Risk Score','Risk Level']]
        for _,row in show.iterrows():
            tdata.append([row['District'],str(row['Year']),row['Month'],row['Predicted_Class'],
                          row['Confidence'],str(row['RiskScore']),row['RiskLevel']])
        dtbl=Table(tdata,colWidths=[3.4*cm,1.4*cm,1.6*cm,2.4*cm,2.2*cm,2.2*cm,2.2*cm],repeatRows=1)
        tstyle=[
            ('BACKGROUND',(0,0),(-1,0),navy),('TEXTCOLOR',(0,0),(-1,0),colors.white),
            ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,-1),7.5),
            ('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('ROWHEIGHT',(0,0),(-1,-1),16),('GRID',(0,0),(-1,-1),0.4,colors.HexColor('#E2E8F0')),
        ]
        risk_cm={'High':red,'Moderate':orange,'Low':green}
        for i in range(1,len(tdata)):
            bg=colors.HexColor('#F7FAFC') if i%2==0 else colors.white
            tstyle.append(('BACKGROUND',(0,i),(-1,i),bg))
            rl=tdata[i][6]
            if rl in risk_cm:
                tstyle.append(('TEXTCOLOR',(6,i),(6,i),risk_cm[rl]))
                tstyle.append(('FONTNAME',(6,i),(6,i),'Helvetica-Bold'))
        dtbl.setStyle(TableStyle(tstyle))
        story.append(dtbl)
        if len(df) > 500:
            story.append(Spacer(1,6))
            story.append(Paragraph(f"<i>Showing first 500 of {len(df)} filtered records. Use the Excel/CSV export for the full dataset.</i>",B))
        story.append(Spacer(1,14))
        story.append(HRFlowable(width="100%",thickness=1,color=colors.HexColor('#E2E8F0'),spaceAfter=8))
        story.append(Paragraph(
            "Dept. of Agriculture &amp; Cooperation, Government of Telangana &nbsp;·&nbsp; District Model v5.0 &nbsp;·&nbsp; Official use only",S))
        doc.build(story)
        return buf.getvalue()
    except ImportError:
        return None


# ─────────────────────────────────────────────────────────────────────────────
#  CHARTS  — each sets its own margin to avoid duplicate-kwarg error
# ─────────────────────────────────────────────────────────────────────────────
def _layout(**kwargs):
    cfg = dict(PLOTLY_BASE)
    cfg['margin'] = kwargs.pop('margin', dict(l=12,r=12,t=42,b=12))
    cfg.update(kwargs)
    return cfg

def chart_class_donut(df):
    """
    Predicted Class Distribution — uses friendlier drought-severity names
    (Normal / Moderate Drought / Severe Drought) mapped from the model's
    Low/Moderate/High output, so this reads as a distinct classification
    view rather than a duplicate of the Risk Level Distribution chart.
    """
    mapped = df['Predicted_Class'].map(DISPLAY_CLASS_LABELS).fillna(df['Predicted_Class'])
    counts = mapped.value_counts().reset_index()
    counts.columns=['Class','Count']
    fig=px.pie(counts,names='Class',values='Count',
               color='Class',color_discrete_map=DISPLAY_CLASS_COLORS,hole=0.50)
    fig.update_traces(textinfo='label+percent',textfont_size=12,
                      pull=[0.04]*len(counts),
                      marker=dict(line=dict(color='white',width=2)))
    fig.update_layout(**_layout(title='Predicted Class Distribution',
                                showlegend=True,
                                legend=dict(orientation='h',y=-0.15,x=0.5,xanchor='center')))
    return fig

def chart_risk_distribution(df):
    counts=df['RiskLevel'].value_counts().reset_index()
    counts.columns=['RiskLevel','Count']
    present=[r for r in RISK_ORDER if r in counts['RiskLevel'].values]
    counts['RiskLevel']=pd.Categorical(counts['RiskLevel'],categories=present,ordered=True)
    counts=counts.sort_values('RiskLevel')
    fig=px.bar(counts,x='RiskLevel',y='Count',color='RiskLevel',
               color_discrete_map=RISK_COLORS,text='Count')
    fig.update_traces(textposition='inside',textfont_size=14,marker_line_width=0)
    fig.update_layout(**_layout(title='Risk Level Distribution',
                                showlegend=False,yaxis_title='Number of District-Months'))
    return fig

def chart_risk_trend(df):
    """Average risk score per period across whatever districts are currently selected."""
    d = df.groupby(['Year','Month_Num','Month','Period'], as_index=False).agg(RiskScore=('RiskScore','mean'))
    d = d.sort_values(['Year','Month_Num'])
    d['RiskLevel'] = d['RiskScore'].apply(lambda s: 'High' if s>=70 else ('Moderate' if s>=40 else 'Low'))
    fig=go.Figure()
    fig.add_trace(go.Scatter(
        x=d['Period'],y=d['RiskScore'],mode='lines+markers',
        line=dict(color='#2B6CB0',width=3),
        marker=dict(size=11,color=d['RiskLevel'].map(RISK_COLORS),
                    line=dict(color='white',width=2)),
        hovertemplate='<b>%{x}</b><br>Avg Risk Score: %{y:.1f}<extra></extra>',
    ))
    fig.add_hrect(y0=70,y1=105,fillcolor='#E53E3E',opacity=0.05,line_width=0)
    fig.add_hrect(y0=40,y1=70,fillcolor='#DD6B20',opacity=0.05,line_width=0)
    fig.update_layout(**_layout(title='Monthly Avg. Risk Score Trend (selected districts)',
                                xaxis_tickangle=-35,yaxis_title='Risk Score (0–100)',
                                yaxis_range=[-2,108]))
    return fig

def chart_confidence_trend(df):
    d = df.groupby(['Year','Month_Num','Month','Period'], as_index=False).agg(Confidence=('Confidence','mean'))
    d = d.sort_values(['Year','Month_Num'])
    pct=(d['Confidence']*100).round(1)
    fig=go.Figure()
    fig.add_trace(go.Scatter(
        x=d['Period'],y=pct,mode='lines+markers',
        line=dict(color='#2C7A7B',width=3,dash='dot'),
        marker=dict(size=9,color='#2C7A7B',line=dict(color='white',width=2)),
        fill='tozeroy',fillcolor='rgba(44,122,123,0.08)',
        hovertemplate='<b>%{x}</b><br>Avg Confidence: %{y:.1f}%<extra></extra>',
    ))
    fig.add_hline(y=70,line_dash='dash',line_color='#DD6B20',line_width=1.5,
                  annotation_text='70% reference',annotation_position='top right',
                  annotation_font=dict(size=10,color='#DD6B20'))
    fig.update_layout(**_layout(title='Model Confidence per Month (%) — selected districts',
                                xaxis_tickangle=-35,yaxis_title='Confidence (%)',
                                yaxis_range=[35,105]))
    return fig

def chart_feature_importance(fi_df):
    norm=fi_df['Importance']/fi_df['Importance'].max()
    clrs=[f"rgba(43,108,176,{0.30+0.70*v:.2f})" for v in norm]
    fig=go.Figure(go.Bar(
        x=fi_df['Importance'],y=fi_df['Feature'],orientation='h',
        marker_color=clrs,marker_line_width=0,
        text=fi_df['Importance'].apply(lambda v:f'{v:.3f}'),textposition='inside',
    ))
    fig.update_layout(**_layout(title='Feature Importance — Correlation with Risk Score',
                                xaxis_title='Normalised Importance',
                                margin=dict(l=170,r=60,t=42,b=12),height=380))
    return fig

def chart_rainfall_vs_risk(df):
    fig=px.scatter(df,x='Rainfall',y='RiskScore',
                   color='RiskLevel',color_discrete_map=RISK_COLORS,
                   size='Confidence',size_max=20,symbol='Predicted_Class',
                   hover_data={'Period':True,'District':True,'Predicted_Class':True,
                               'Confidence':':.2f','Rainfall':':.1f','RiskScore':':.1f'},
                   labels={'Rainfall':'Rainfall (mm)','RiskScore':'Risk Score'})
    fig.update_layout(**_layout(title='Rainfall vs Risk Score'))
    return fig

def chart_spi_trend(df):
    d = df.groupby(['Year','Month_Num','Month','Period'], as_index=False).agg(SPI3=('SPI3','mean'))
    d = d.sort_values(['Year','Month_Num'])
    clrs=d['SPI3'].apply(lambda v:'#E53E3E' if v<0 else '#38A169')
    fig=go.Figure(go.Bar(
        x=d['Period'],y=d['SPI3'],marker_color=clrs,marker_line_width=0,
        hovertemplate='<b>%{x}</b><br>Avg SPI-3: %{y:.3f}<extra></extra>',
    ))
    fig.add_hline(y=0,line_color='#2D3748',line_width=1.5)
    fig.add_hline(y=-1,line_dash='dash',line_color='#E53E3E',line_width=1.2,
                  annotation_text='Drought threshold (−1)',
                  annotation_position='bottom right',
                  annotation_font=dict(size=10,color='#E53E3E'))
    fig.update_layout(**_layout(title='SPI-3 Monthly Index (avg. of selected districts)',
                                xaxis_tickangle=-35,yaxis_title='SPI-3'))
    return fig

def _normalize_name(s: str) -> str:
    """Uppercase and strip everything except letters, so 'Medchal-Malkajgiri',
    'Medchal Malkajgiri' and 'MEDCHAL MALKAJGIRI' all collapse to the same key."""
    return re.sub(r'[^A-Z]', '', s.upper())

CANONICAL_BY_NORM = {_normalize_name(k): k for k in DISTRICTS}
GEOJSON_NAME_MAP_NORM = {_normalize_name(k): v for k, v in GEOJSON_NAME_MAP.items()}

@st.cache_data(show_spinner=False)
def load_district_geojson(path: str):
    """Loads the Telangana district polygon GeoJSON and normalizes the district
    name onto a 'District' property so it always matches the CSV's District column.
    Returns (geojson_dict_or_None, list_of_unmatched_raw_names)."""
    if not os.path.exists(path):
        return None, []
    with open(path, "r", encoding="utf-8") as f:
        gj = json.load(f)
    unmatched = []
    for feat in gj.get("features", []):
        props = feat.setdefault("properties", {})
        raw = None
        for k in GEOJSON_NAME_KEYS:
            if k in props and props[k]:
                raw = str(props[k]); break
        if raw is None:
            raw = str(next(iter(props.values()), "Unknown"))
        norm = _normalize_name(raw)
        canonical = CANONICAL_BY_NORM.get(norm) or GEOJSON_NAME_MAP_NORM.get(norm)
        if canonical is None:
            close = difflib.get_close_matches(norm, CANONICAL_BY_NORM.keys(), n=1, cutoff=0.72)
            if close:
                canonical = CANONICAL_BY_NORM[close[0]]
        if canonical is None:
            canonical = raw.strip()
            unmatched.append(raw.strip())
        props["District"] = canonical
    return gj, unmatched


def build_district_lookup(dist_df, month_filter=None, year_filter=None):
    """One row per district — averaged over whatever Year/Month filter is active.
    Feeds both the polygon fill color and the hover tooltip fields."""
    d = dist_df.copy()
    if year_filter:  d = d[d["Year"] == year_filter]
    if month_filter: d = d[d["Month"] == month_filter]
    if d.empty:
        return pd.DataFrame()

    agg_cols = {c: "mean" for c in
                ["Rainfall", "Temperature", "Soil_Moisture", "NDVI", "SPI3",
                 "RiskScore", "Confidence", "Groundwater_Proxy"]
                if c in d.columns}
    out = d.groupby("District", as_index=False).agg(agg_cols).round(3)

    # Most frequent categorical fields per district in the filtered window
    for cat_col in ["RiskLevel", "Predicted_Class"]:
        if cat_col in d.columns:
            mode_map = d.groupby("District")[cat_col].agg(
                lambda s: s.value_counts().idxmax())
            out[cat_col] = out["District"].map(mode_map)
    return out


def build_gis_map(lookup_df, geojson_data, basemap="Street", color_field="RiskLevel",
                   color_mode="discrete", selected_district=None, height=560):
    m = folium.Map(location=[17.9, 79.2], zoom_start=7, tiles=None, control_scale=True)

    # Register every basemap as its own toggleable Leaflet layer (radio buttons in the
    # layer control, top-right) instead of baking in a single fixed tile source. Only the
    # one chosen in the sidebar is shown by default; the others are one click away.
    for name, cfg in BASEMAP_TILES.items():
        folium.TileLayer(
            tiles=cfg["tiles"], attr=cfg["attr"], name=name,
            show=(name == basemap), control=True, overlay=False,
        ).add_to(m)

    if geojson_data is None or lookup_df.empty:
        return m

    row_by_district = lookup_df.set_index("District").to_dict("index")

    if color_mode == "discrete":
        def get_color(dist_name):
            row = row_by_district.get(dist_name)
            if not row: return "#CBD5E0"
            return RISK_COLORS.get(row.get(color_field), "#CBD5E0")
    else:
        vals = lookup_df[color_field].dropna()
        vmin, vmax = (float(vals.min()), float(vals.max())) if len(vals) else (0, 1)
        cmap = bcm.LinearColormap(["#38A169", "#DD6B20", "#E53E3E"], vmin=vmin, vmax=vmax)
        def get_color(dist_name):
            row = row_by_district.get(dist_name)
            if not row or pd.isna(row.get(color_field)): return "#CBD5E0"
            return cmap(row[color_field])

    def style_fn(feature):
        dname = feature["properties"]["District"]
        is_sel = selected_district and dname == selected_district
        return {
            "fillColor": get_color(dname),
            "color": "#FFD700" if is_sel else "#1A365D",
            "weight": 3 if is_sel else 1,
            "fillOpacity": 0.85 if is_sel else 0.65,
        }

    def highlight_fn(feature):
        return {"weight": 3, "color": "#2B6CB0", "fillOpacity": 0.9}

    tooltip_fields, tooltip_aliases = ["District"], ["District:"]
    field_alias_map = [
        ("Rainfall", "Rainfall (mm):"), ("Temperature", "Temperature (°C):"),
        ("Soil_Moisture", "Soil Moisture:"), ("NDVI", "NDVI:"), ("SPI3", "SPI-3:"),
        ("Predicted_Class", "Prediction:"), ("Confidence", "Confidence:"),
        ("RiskScore", "Risk Score:"), ("Groundwater_Proxy", "Groundwater Proxy:"),
    ]
    sample_props = next(iter(row_by_district.values()), {})
    for col, alias in field_alias_map:
        if col in sample_props:
            tooltip_fields.append(col); tooltip_aliases.append(alias)

    # Attach the per-district lookup values onto each polygon's properties so the
    # tooltip can read them directly off the feature.
    for feat in geojson_data["features"]:
        dname = feat["properties"]["District"]
        feat["properties"].update(row_by_district.get(dname, {}))

    gj_layer = folium.GeoJson(
        geojson_data, name="Prediction Layer",
        style_function=style_fn, highlight_function=highlight_fn,
        tooltip=folium.GeoJsonTooltip(fields=tooltip_fields, aliases=tooltip_aliases,
                                       sticky=True, localize=True),
    ).add_to(m)

    Search(layer=gj_layer, geom_type="Polygon", search_label="District",
           placeholder="Search district…", collapsed=False,
           position="topright").add_to(m)

    Fullscreen(position="topleft").add_to(m)
    MiniMap(toggle_display=True, position="bottomright").add_to(m)
    LocateControl(position="topleft").add_to(m)
    MeasureControl(primary_length_unit="kilometers", position="bottomleft").add_to(m)
    folium.LayerControl(position="topright", collapsed=False).add_to(m)

    return m


def chart_district_bar(dist_df, month_filter=None, year_filter=None):
    d=dist_df.copy()
    if year_filter:  d=d[d['Year']==year_filter]
    if month_filter: d=d[d['Month']==month_filter]
    if d.empty: return None
    agg=d.groupby('District',as_index=False)['RiskScore'].mean().round(1)
    agg=agg.sort_values('RiskScore',ascending=True)
    agg['Color']=agg['RiskScore'].apply(
        lambda v:'#E53E3E' if v>=70 else ('#DD6B20' if v>=40 else '#38A169'))
    fig=go.Figure(go.Bar(
        x=agg['RiskScore'],y=agg['District'],orientation='h',
        marker_color=agg['Color'],marker_line_width=0,
        text=agg['RiskScore'].astype(str),textposition='inside',
        hovertemplate='<b>%{y}</b><br>Avg Risk Score: %{x}<extra></extra>',
    ))
    fig.update_layout(**_layout(title='Average Risk Score by District',
                                xaxis_title='Risk Score (0–100)',
                                margin=dict(l=140,r=60,t=42,b=12),
                                height=700,
                                yaxis=dict(tickfont=dict(size=10))))
    return fig

def chart_shap_waterfall(shap_df: pd.DataFrame, base_value: float, pred_class: str):
    d = shap_df.sort_values('SHAP_Value').copy()
    colors = d['SHAP_Value'].apply(lambda v: '#E53E3E' if v > 0 else '#38A169')
    labels = d.apply(lambda r: f"{r['Feature']} = {r['Value']:.3f}", axis=1)

    fig = go.Figure(go.Bar(
        x=d['SHAP_Value'], y=labels, orientation='h',
        marker_color=colors, marker_line_width=0,
        text=d['SHAP_Value'].apply(lambda v: f"{v:+.3f}"),
        textposition='inside', textfont=dict(color='white', size=11),
        hovertemplate='<b>%{y}</b><br>SHAP impact: %{x:+.4f}<extra></extra>',
    ))
    fig.add_vline(x=0, line_color='#1A202C', line_width=1.5)
    fig.update_layout(**_layout(
        title=f'Feature Contributions Toward "{pred_class}" Prediction',
        xaxis_title='SHAP Value (impact on prediction)',
        height=320 + 20*len(d),
        margin=dict(l=200, r=40, t=42, b=12),
    ))
    return fig


def show_dashboard():
    inject_css()

    if not os.path.exists(DISTRICT_CSV_PATH):
        st.error(f"❌ CSV not found: `{DISTRICT_CSV_PATH}` — place the output of "
                 f"`train_drought_model.py` (District_Drought_Predictions.csv) in the same folder as this script.")
        st.stop()

    with st.spinner("Loading district data…"):
        df_raw = load_district_predictions(DISTRICT_CSV_PATH)

    years_all     = sorted(df_raw['Year'].unique().tolist())
    months_all    = [m for m in MONTH_ORDER if m in df_raw['Month'].unique()]
    risks_all     = [r for r in RISK_ORDER   if r in df_raw['RiskLevel'].unique()]
    classes_all   = [c for c in RISK_ORDER   if c in df_raw['Predicted_Class'].unique()]
    districts_all = sorted(df_raw['District'].unique().tolist())

    # ── SIDEBAR ───────────────────────────────────────────────────────────────
    with st.sidebar:
        uname    = st.session_state.get("user_name","User")
        urole    = st.session_state.get("user_role","—")
        lt       = st.session_state.get("login_time","")
        initials = "".join(w[0].upper() for w in uname.split()[:2])

        st.markdown(f"""
        <div class="sb-header">
            <span class="sb-icon">🌾</span>
            <p class="sb-title">Telangana Drought<br>Prediction System</p>
            <p class="sb-sub">Dept. of Agriculture · GoT · District Model v5.0</p>
        </div>
        <div class="sb-user">
            <div class="sb-av">{initials}</div>
            <div>
                <div class="sb-uname">{uname}</div>
                <div class="sb-urole">{urole} · {lt}</div>
            </div>
        </div>""", unsafe_allow_html=True)

        st.markdown('<div class="sb-sec">📊 Dashboard Filters</div>', unsafe_allow_html=True)
        sel_years     = st.multiselect("Year",           options=years_all,     default=[])
        sel_months    = st.multiselect("Month",          options=months_all,    default=[])
        sel_districts = st.multiselect("District",       options=districts_all, default=[],
                                       key="sel_districts")
        sel_risks     = st.multiselect("Risk Level",     options=risks_all,     default=[])
        sel_classes   = st.multiselect("Predicted Class",options=classes_all,   default=[],
                                       format_func=lambda c: DISPLAY_CLASS_LABELS.get(c,c))

        st.markdown('<div class="sb-sec">🔍 Prediction Lookup</div>', unsafe_allow_html=True)
        lu_year     = st.selectbox("Lookup Year",     options=years_all)
        lu_month    = st.selectbox("Lookup Month",    options=months_all)
        lu_district = st.selectbox("Lookup District", options=districts_all, key="lu_district")

        st.markdown('<div class="sb-sec">🗺️ GIS Map Controls</div>', unsafe_allow_html=True)
        basemap_choice = st.selectbox("Basemap", options=list(BASEMAP_TILES.keys()),
                                       key="basemap_choice")
        map_layer_label = st.selectbox("Parameters", options=list(MAP_INDICATOR_OPTIONS.keys()),
                                        key="map_layer_label")
        map_year = (st.select_slider("Year", options=years_all, value=years_all[-1],
                                      key="map_year")
                    if years_all else None)
        map_month    = st.selectbox("Map Month",    options=['All']+months_all, key="map_month")
        map_district = st.selectbox("Zoom District", options=['All']+districts_all,
                                     key="map_district")

        st.markdown('<hr class="sb-divider">', unsafe_allow_html=True)
        if st.button("🚪 Logout", use_container_width=True):
            for k in ["logged_in","username","user_name","user_role","login_time"]:
                st.session_state.pop(k,None)
            st.rerun()

        st.markdown("""
        <div style='font-size:0.62rem;color:rgba(190,227,248,0.30);
                    line-height:1.8;margin-top:12px;padding:0 4px;'>
        Sources: IMD · ISRO · CWC · CGWB · NRSC<br>
        Model: District Stacking Ensemble (RF+XGBoost→LogReg)<br>
        Indicators: SPI-3 · NDVI · Soil Moisture · Groundwater
        </div>""", unsafe_allow_html=True)

    # ── FILTERS ───────────────────────────────────────────────────────────────
    df=df_raw.copy()
    if sel_years:     df=df[df['Year'].isin(sel_years)]
    if sel_months:    df=df[df['Month'].isin(sel_months)]
    if sel_districts: df=df[df['District'].isin(sel_districts)]
    if sel_risks:     df=df[df['RiskLevel'].isin(sel_risks)]
    if sel_classes:   df=df[df['Predicted_Class'].isin(sel_classes)]

    if df.empty:
        st.warning("⚠️ No records match the selected filters.")
        return

    # ── HEADER ────────────────────────────────────────────────────────────────
    st.markdown(f"""
    <div class="gov-header">
        <span class="gov-badge">Government of Telangana · Department of Agriculture &amp; Cooperation</span>
        <h1>🌾 Telangana AI-Based District Drought Prediction &amp; Risk Assessment System</h1>
        <p>District-Level AI/ML Early-Warning Platform · SPI-3 · NDVI · Soil Moisture · Groundwater
           &nbsp;·&nbsp; <strong>{df['District'].nunique()}</strong> districts
           &nbsp;·&nbsp; Showing <strong>{len(df)}</strong> of <strong>{len(df_raw)}</strong> district-month records</p>
    </div>""", unsafe_allow_html=True)

    # ── KPI CARDS ─────────────────────────────────────────────────────────────
    high_n    = int((df['Predicted_Class']=='High').sum())
    low_n     = int((df['Predicted_Class']=='Low').sum())
    avg_conf  = df['Confidence'].mean()*100
    avg_risk  = df['RiskScore'].mean()
    high_pct  = (df['RiskLevel']=='High').mean()*100

    c1,c2,c3,c4=st.columns(4)
    with c1:
        st.markdown(f"""<div class="kpi-card red">
            <div class="kpi-label">🔴 High Risk Records</div>
            <div class="kpi-value">{high_n}</div>
            <div class="kpi-sub">of {len(df)} filtered district-months</div>
        </div>""", unsafe_allow_html=True)
    with c2:
        st.markdown(f"""<div class="kpi-card green">
            <div class="kpi-label">🟢 Low Risk Records</div>
            <div class="kpi-value">{low_n}</div>
            <div class="kpi-sub">Normal / adequate conditions</div>
        </div>""", unsafe_allow_html=True)
    with c3:
        st.markdown(f"""<div class="kpi-card teal">
            <div class="kpi-label">🎯 Avg. Confidence</div>
            <div class="kpi-value">{avg_conf:.1f}%</div>
            <div class="kpi-sub">Model prediction confidence</div>
        </div>""", unsafe_allow_html=True)
    with c4:
        card_cls='red' if avg_risk>=70 else ('orange' if avg_risk>=40 else 'teal')
        st.markdown(f"""<div class="kpi-card {card_cls}">
            <div class="kpi-label">⚡ Avg. Risk Score</div>
            <div class="kpi-value">{avg_risk:.1f}</div>
            <div class="kpi-sub">High-risk share: {high_pct:.0f}%</div>
        </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ── TABS ──────────────────────────────────────────────────────────────────
    # Lazy render: only the ACTIVE tab's code runs each rerun. st.tabs() runs every
    # tab's body on every single interaction anywhere in the app (that's what was
    # causing the lag/unresponsive tab-switching) — this radio-based nav only runs
    # the one tab you're actually looking at.
    TAB_LABELS = [
        "📋 Overview","📈 Charts","📊 Year Comparison","📅 Seasonal Analysis",
        "🌡️ Indicator Deep Dive","🏆 Month Rankings","🗺️ District Map",
        "🔍 Prediction Lookup","📋 Advisory Board","🤖 Live Prediction",
        "🧬 Why This Prediction?","📄 Reports & Export"
    ]
    if "active_main_tab" not in st.session_state:
        st.session_state["active_main_tab"] = TAB_LABELS[0]

    active_tab = st.radio(
        "Navigation", TAB_LABELS,
        key="active_main_tab",
        horizontal=True,
        label_visibility="collapsed",
    )

    # ════════ TAB 1 — Overview ════════
    if active_tab == TAB_LABELS[0]:
        col_tbl,col_pie=st.columns([3,2])
        with col_tbl:
            st.markdown('<div class="section-card">',unsafe_allow_html=True)
            st.markdown('<p class="section-title">📋 Prediction Summary Table</p>',unsafe_allow_html=True)
            show_df=df[['District','Year','Month','Month_Num','Predicted_Class','Confidence',
                         'RiskScore','RiskLevel','Rainfall','Temperature','SPI3','NDVI']].copy()
            show_df=show_df.sort_values(['District','Year','Month_Num']).drop(columns=['Month_Num'])
            show_df['Confidence']=(show_df['Confidence']*100).round(1).astype(str)+'%'
            for col in ['RiskScore','Rainfall','Temperature']:
                show_df[col]=show_df[col].round(1)
            show_df['SPI3']=show_df['SPI3'].round(3)
            show_df['NDVI']=show_df['NDVI'].round(3)
            show_df.columns=['District','Year','Month','Predicted Class','Confidence',
                             'Risk Score','Risk Level','Rainfall (mm)','Temp (°C)','SPI-3','NDVI']
            st.dataframe(show_df.reset_index(drop=True),
                         use_container_width=True,height=320,hide_index=True)
            st.markdown('</div>',unsafe_allow_html=True)
        with col_pie:
            st.markdown('<div class="section-card">',unsafe_allow_html=True)
            st.markdown('<p class="section-title">🥧 Class Distribution</p>',unsafe_allow_html=True)
            st.plotly_chart(chart_class_donut(df),use_container_width=True,
                            config={'displayModeBar':False})
            st.markdown('</div>',unsafe_allow_html=True)

        col_rd,col_sp=st.columns(2)
        with col_rd:
            st.markdown('<div class="section-card">',unsafe_allow_html=True)
            st.markdown('<p class="section-title">📊 Risk Level Distribution</p>',unsafe_allow_html=True)
            st.plotly_chart(chart_risk_distribution(df),use_container_width=True,
                            config={'displayModeBar':False})
            st.markdown('</div>',unsafe_allow_html=True)
        with col_sp:
            st.markdown('<div class="section-card">',unsafe_allow_html=True)
            st.markdown('<p class="section-title">📉 SPI-3 Index Timeline</p>',unsafe_allow_html=True)
            st.plotly_chart(chart_spi_trend(df),use_container_width=True,
                            config={'displayModeBar':False})
            st.markdown('</div>',unsafe_allow_html=True)

    # ════════ TAB 2 — Charts ════════
    if active_tab == TAB_LABELS[1]:
        col_rt,col_ct=st.columns(2)
        with col_rt:
            st.markdown('<div class="section-card">',unsafe_allow_html=True)
            st.markdown('<p class="section-title">📈 Monthly Risk Score Trend</p>',unsafe_allow_html=True)
            st.plotly_chart(chart_risk_trend(df),use_container_width=True,
                            config={'displayModeBar':False})
            st.markdown('</div>',unsafe_allow_html=True)
        with col_ct:
            st.markdown('<div class="section-card">',unsafe_allow_html=True)
            st.markdown('<p class="section-title">🎯 Model Confidence Trend</p>',unsafe_allow_html=True)
            st.plotly_chart(chart_confidence_trend(df),use_container_width=True,
                            config={'displayModeBar':False})
            st.markdown('</div>',unsafe_allow_html=True)

        col_fi,col_sc=st.columns(2)
        with col_fi:
            fi_df=compute_feature_importance(df)
            if not fi_df.empty:
                st.markdown('<div class="section-card">',unsafe_allow_html=True)
                st.markdown('<p class="section-title">🧠 Feature Importance</p>',unsafe_allow_html=True)
                st.plotly_chart(chart_feature_importance(fi_df),use_container_width=True,
                                config={'displayModeBar':False})
                st.markdown('</div>',unsafe_allow_html=True)
        with col_sc:
            st.markdown('<div class="section-card">',unsafe_allow_html=True)
            st.markdown('<p class="section-title">🌧️ Rainfall vs Risk Score</p>',unsafe_allow_html=True)
            st.plotly_chart(chart_rainfall_vs_risk(df),use_container_width=True,
                            config={'displayModeBar':False})
            st.markdown('</div>',unsafe_allow_html=True)

    # ════════ TAB 7 — District Map (REAL data) ════════
    if active_tab == TAB_LABELS[6]:
        mf = None if map_month=='All' else map_month
        indicator_label = st.session_state.get("map_layer_label", "Prediction Risk")
        color_field, color_mode = MAP_INDICATOR_OPTIONS[indicator_label]

        lookup_df = build_district_lookup(df_raw, month_filter=mf, year_filter=map_year)
        geojson_data, unmatched_names = load_district_geojson(DISTRICT_GEOJSON_PATH)

        st.markdown('<div class="section-card">', unsafe_allow_html=True)
        st.markdown('<p class="section-title">🗺️ Telangana District Drought Risk Map — WebGIS</p>',
                    unsafe_allow_html=True)
        st.caption("Click any district to drill down across Overview, Charts, Advisory, "
                   "Live Prediction and SHAP · Hover for full indicator values · "
                   f"Coloured by: {indicator_label}")

        if geojson_data is None:
            st.error(f"District boundary file not found at `{DISTRICT_GEOJSON_PATH}`. "
                     "Add a Telangana district GeoJSON next to the script to enable the "
                     "polygon map (see setup notes).")
        elif lookup_df.empty:
            st.info("No district data for the selected map filters.")
        else:
            if unmatched_names:
                st.warning(
                    f"⚠️ {len(set(unmatched_names))} polygon(s) in the GeoJSON didn't match a "
                    f"known district name and will render gray (no data): "
                    f"**{', '.join(sorted(set(unmatched_names)))}**. Open the GeoJSON, check the "
                    "spelling for these, and add an entry to `GEOJSON_NAME_MAP` near the top of "
                    "the script mapping it to the matching name in `DISTRICTS`.")
            active_district = st.session_state.get("selected_district") \
                               if map_district == 'All' else map_district

            gis_map = build_gis_map(
                lookup_df, geojson_data,
                basemap=st.session_state.get("basemap_choice", "Street"),
                color_field=color_field, color_mode=color_mode,
                selected_district=active_district,
            )

            # Legend
            if color_mode == "discrete":
                st.markdown("""
                <div style="display:flex;gap:18px;margin:4px 0 10px;font-size:0.85rem;">
                    <span>🟢 Low</span><span>🟡 Moderate</span><span>🔴 High</span>
                </div>""", unsafe_allow_html=True)

            map_data = st_folium(gis_map, height=560, use_container_width=True, key="gis_map")

            clicked = None
            if map_data and map_data.get("last_active_drawing"):
                clicked = map_data["last_active_drawing"]["properties"].get("District")

            if clicked and clicked != st.session_state.get("selected_district"):
                st.session_state["selected_district"]  = clicked
                st.session_state["sel_districts"]       = [clicked]
                st.session_state["map_district"]        = clicked
                st.session_state["lu_district"]         = clicked
                st.session_state["manual_ref_district"] = clicked
                st.session_state["shap_district"]       = clicked
                st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)

        st.markdown('<div class="section-card">', unsafe_allow_html=True)
        st.markdown('<p class="section-title">📊 District Risk Ranking</p>', unsafe_allow_html=True)
        bar_fig = chart_district_bar(df_raw, month_filter=mf, year_filter=map_year)
        if bar_fig:
            st.plotly_chart(bar_fig, use_container_width=True, config={'displayModeBar':False})
        st.markdown('</div>', unsafe_allow_html=True)

        if map_district != 'All':
            d_df = df_raw[df_raw['District']==map_district].sort_values(['Year','Month_Num'])

            if d_df.empty:
                st.info(f"No data found for {map_district}.")
            else:
                st.markdown(f'<div class="section-card">', unsafe_allow_html=True)
                st.markdown(f'<p class="section-title">📍 District Profile — {map_district}</p>',
                            unsafe_allow_html=True)

                avg_risk_d = d_df['RiskScore'].mean()
                avg_rain  = d_df['Rainfall'].mean()
                avg_sm    = d_df['Soil_Moisture'].mean() if 'Soil_Moisture' in d_df.columns else None
                avg_ndvi  = d_df['NDVI'].mean()          if 'NDVI'          in d_df.columns else None
                high_n_d  = int((d_df['RiskLevel']=='High').sum())
                lat_val   = d_df['Lat'].iloc[0]
                lon_val   = d_df['Lon'].iloc[0]

                ka,kb,kc,kd = st.columns(4)
                with ka:
                    st.markdown(f"""<div class="kpi-card {'rust' if avg_risk_d>=70 else 'amber' if avg_risk_d>=40 else 'green'}">
                        <div class="kpi-label">⚡ Avg Risk Score</div>
                        <div class="kpi-value">{avg_risk_d:.1f}</div>
                        <div class="kpi-sub">All available periods</div>
                    </div>""", unsafe_allow_html=True)
                with kb:
                    st.markdown(f"""<div class="kpi-card rust">
                        <div class="kpi-label">🔴 High Risk Months</div>
                        <div class="kpi-value">{high_n_d}</div>
                        <div class="kpi-sub">out of {len(d_df)} records</div>
                    </div>""", unsafe_allow_html=True)
                with kc:
                    st.markdown(f"""<div class="kpi-card teal">
                        <div class="kpi-label">🌧️ Avg Rainfall</div>
                        <div class="kpi-value">{avg_rain:.1f}</div>
                        <div class="kpi-sub">mm per month</div>
                    </div>""", unsafe_allow_html=True)
                with kd:
                    ndvi_display = f"{avg_ndvi:.3f}" if avg_ndvi is not None else "N/A"
                    st.markdown(f"""<div class="kpi-card green">
                        <div class="kpi-label">🌿 Avg NDVI</div>
                        <div class="kpi-value">{ndvi_display}</div>
                        <div class="kpi-sub">Vegetation health index</div>
                    </div>""", unsafe_allow_html=True)

                st.markdown("<br>", unsafe_allow_html=True)

                # ── Zoomed GIS map ──
                st.markdown("#### 🔍 Zoomed District View")
                zoom_lookup = build_district_lookup(df_raw, month_filter=mf, year_filter=map_year)
                if geojson_data is not None and not zoom_lookup.empty:
                    zoom_map = build_gis_map(
                        zoom_lookup, geojson_data,
                        basemap=st.session_state.get("basemap_choice", "Street"),
                        color_field=color_field, color_mode=color_mode,
                        selected_district=map_district,
                    )
                    zoom_map.location = [float(lat_val), float(lon_val)]
                    zoom_map.zoom_start = 9
                    st_folium(zoom_map, height=460, use_container_width=True,
                              key="gis_zoom_map")

                # ── Monthly trend charts ──
                st.markdown("#### 📈 Monthly Trends for this District")

                col_r1, col_r2 = st.columns(2)
                with col_r1:
                    fig_rs = go.Figure(go.Scatter(
                        x=d_df['Period'], y=d_df['RiskScore'],
                        mode='lines+markers',
                        line=dict(color='#E53E3E', width=2.5),
                        marker=dict(size=9, color=d_df['RiskLevel'].map(RISK_COLORS),
                                    line=dict(color='white', width=1.5)),
                        hovertemplate='<b>%{x}</b><br>Risk Score: %{y:.1f}<extra></extra>',
                    ))
                    fig_rs.update_layout(**_layout(
                        title=f'Risk Score — {map_district}',
                        xaxis_tickangle=-35, yaxis_title='Risk Score',
                        yaxis_range=[0, 105],
                    ))
                    st.plotly_chart(fig_rs, use_container_width=True,
                                    config={'displayModeBar':False})

                with col_r2:
                    fig_rf = go.Figure(go.Bar(
                        x=d_df['Period'], y=d_df['Rainfall'],
                        marker_color='#3182CE', marker_line_width=0,
                        hovertemplate='<b>%{x}</b><br>Rainfall: %{y:.1f} mm<extra></extra>',
                    ))
                    fig_rf.update_layout(**_layout(
                        title=f'Rainfall (mm) — {map_district}',
                        xaxis_tickangle=-35, yaxis_title='Rainfall (mm)',
                    ))
                    st.plotly_chart(fig_rf, use_container_width=True,
                                    config={'displayModeBar':False})

                col_r3, col_r4 = st.columns(2)
                with col_r3:
                    if 'NDVI' in d_df.columns:
                        fig_ndvi = go.Figure(go.Scatter(
                            x=d_df['Period'], y=d_df['NDVI'],
                            mode='lines+markers', fill='tozeroy',
                            line=dict(color='#38A169', width=2.5),
                            marker=dict(size=7, color='#38A169'),
                            hovertemplate='<b>%{x}</b><br>NDVI: %{y:.4f}<extra></extra>',
                        ))
                        fig_ndvi.update_layout(**_layout(
                            title=f'NDVI Index — {map_district}',
                            xaxis_tickangle=-35, yaxis_title='NDVI',
                            yaxis_range=[0, 1],
                        ))
                        st.plotly_chart(fig_ndvi, use_container_width=True,
                                        config={'displayModeBar':False})

                with col_r4:
                    if 'Soil_Moisture' in d_df.columns:
                        fig_sm = go.Figure(go.Scatter(
                            x=d_df['Period'], y=d_df['Soil_Moisture'],
                            mode='lines+markers', fill='tozeroy',
                            line=dict(color='#744210', width=2.5),
                            marker=dict(size=7, color='#744210'),
                            hovertemplate='<b>%{x}</b><br>Soil Moisture: %{y:.4f}<extra></extra>',
                        ))
                        fig_sm.update_layout(**_layout(
                            title=f'Soil Moisture — {map_district}',
                            xaxis_tickangle=-35, yaxis_title='Soil Moisture',
                            yaxis_range=[0, 1],
                        ))
                        st.plotly_chart(fig_sm, use_container_width=True,
                                        config={'displayModeBar':False})

                # ── Monthly data table ──
                st.markdown("#### 📋 Monthly Data Table")
                show_d = d_df[['Year','Month','RiskScore','RiskLevel','Predicted_Class',
                                'Rainfall','Temperature','NDVI','Soil_Moisture']].copy()
                show_d['RiskScore']     = show_d['RiskScore'].round(1)
                show_d['Rainfall']      = show_d['Rainfall'].round(1)
                show_d['Temperature']   = show_d['Temperature'].round(1)
                show_d['NDVI']          = show_d['NDVI'].round(4)
                show_d['Soil_Moisture'] = show_d['Soil_Moisture'].round(4)
                show_d.columns = ['Year','Month','Risk Score','Risk Level','Predicted Class',
                                   'Rainfall (mm)','Temp (°C)','NDVI','Soil Moisture']
                st.dataframe(show_d.reset_index(drop=True),
                             use_container_width=True, height=300, hide_index=True)

                st.markdown('</div>', unsafe_allow_html=True)

    # ════════ TAB 8 — Prediction Lookup (Year + Month + District) ════════
    if active_tab == TAB_LABELS[7]:
        st.markdown('<div class="section-card">',unsafe_allow_html=True)
        st.markdown('<p class="section-title">🔍 Point Prediction Lookup</p>',unsafe_allow_html=True)
        lu_row=df_raw[(df_raw['Year']==lu_year)&(df_raw['Month']==lu_month)&(df_raw['District']==lu_district)]
        if lu_row.empty:
            st.info(f"No record found for **{lu_district} — {lu_month} {lu_year}**.")
        else:
            r          = lu_row.iloc[0]
            cls        = r['Predicted_Class']
            conf_pct   = r['Confidence']*100
            risk_s     = r['RiskScore']
            risk_l     = r['RiskLevel']
            advisory   = str(r['Advisory']).replace('\\n',' ').replace('\n',' ').strip()
            cls_color  = CLASS_COLORS.get(cls,'#1A365D')
            risk_color = RISK_COLORS.get(risk_l,'#1A365D')
            alert_cls  = 'alert-high' if risk_l=='High' else ('alert-mod' if risk_l=='Moderate' else 'alert-low')
            icon       = '🔴' if risk_l=='High' else ('🟠' if risk_l=='Moderate' else '🟢')
            st.markdown(f"""
            <div class="{alert_cls}">
                <strong>{icon} {risk_l} Risk — {lu_district} — {MONTH_FULL.get(lu_month,lu_month)} {lu_year}</strong><br>
                <span style='font-size:0.84rem;'>{advisory}</span>
            </div>""", unsafe_allow_html=True)
            col_l2,col_r2=st.columns(2)
            with col_l2:
                st.markdown(f"""
                <div class="lookup-box">
                    <h4>📍 {lu_district} — {MONTH_FULL.get(lu_month,lu_month)} {lu_year}</h4>
                    <div class="lrow"><span class="lkey">Predicted Class</span>
                        <span class="lval" style="color:{cls_color}">● {cls}</span></div>
                    <div class="lrow"><span class="lkey">Model Confidence</span>
                        <span class="lval">{conf_pct:.2f}%</span></div>
                    <div class="lrow"><span class="lkey">Risk Score</span>
                        <span class="lval">{risk_s:.1f} / 100</span></div>
                    <div class="lrow"><span class="lkey">Risk Level</span>
                        <span class="lval" style="color:{risk_color}">▲ {risk_l}</span></div>
                </div>""", unsafe_allow_html=True)
            with col_r2:
                gw_val = r['Groundwater_Proxy'] if 'Groundwater_Proxy' in r else None
                gw_txt = f"{gw_val:.3f}" if gw_val is not None else "N/A"
                st.markdown(f"""
                <div class="lookup-box">
                    <h4>🌦️ Meteorological Indicators</h4>
                    <div class="lrow"><span class="lkey">Rainfall</span>
                        <span class="lval">{r['Rainfall']:.1f} mm</span></div>
                    <div class="lrow"><span class="lkey">Temperature</span>
                        <span class="lval">{r['Temperature']:.1f} °C</span></div>
                    <div class="lrow"><span class="lkey">SPI-3 Index</span>
                        <span class="lval">{r['SPI3']:.4f}</span></div>
                    <div class="lrow"><span class="lkey">NDVI</span>
                        <span class="lval">{r['NDVI']:.4f}</span></div>
                    <div class="lrow"><span class="lkey">Soil Moisture</span>
                        <span class="lval">{r['Soil_Moisture']:.2f}</span></div>
                    <div class="lrow"><span class="lkey">Groundwater Proxy</span>
                        <span class="lval">{gw_txt}</span></div>
                </div>""", unsafe_allow_html=True)
        st.markdown('</div>',unsafe_allow_html=True)

    # ════════ TAB 12 — Reports & Export ════════
    if active_tab == TAB_LABELS[11]:
        st.markdown('<div class="section-card">',unsafe_allow_html=True)
        st.markdown('<p class="section-title">📄 Download Reports &amp; Data</p>',unsafe_allow_html=True)
        col_e1,col_e2,col_e3=st.columns(3)
        with col_e1:
            st.markdown("#### 📊 Excel Export")
            st.caption("Filtered data with styled headers")
            st.download_button(
                label="⬇️ Download Excel (.xlsx)",
                data=to_excel_bytes(df),
                file_name=f"Telangana_District_Drought_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)
        with col_e2:
            st.markdown("#### 📋 PDF Report")
            st.caption("Formatted official report — install reportlab first")
            if st.button("🔄 Generate PDF",use_container_width=True):
                with st.spinner("Building PDF…"):
                    pdf_bytes=generate_pdf_report(df,sel_years,sel_months,sel_districts,
                                                   st.session_state.get("user_name","User"))
                if pdf_bytes:
                    st.download_button(
                        label="⬇️ Download PDF",data=pdf_bytes,
                        file_name=f"Telangana_District_Drought_{datetime.datetime.now().strftime('%Y%m%d')}.pdf",
                        mime="application/pdf",use_container_width=True)
                else:
                    st.warning("Run: `pip install reportlab` then try again.")
        with col_e3:
            st.markdown("#### 📥 CSV Export")
            st.caption("Raw filtered data as CSV")
            st.download_button(
                label="⬇️ Download CSV",
                data=df.to_csv(index=False).encode('utf-8'),
                file_name=f"Telangana_District_Drought_{datetime.datetime.now().strftime('%Y%m%d')}.csv",
                mime="text/csv",use_container_width=True)
        st.markdown('</div>',unsafe_allow_html=True)

    # ════════ TAB 3 — Year Comparison ════════
    if active_tab == TAB_LABELS[2]:
        st.markdown('<div class="section-card">',unsafe_allow_html=True)
        st.markdown('<p class="section-title">📊 Year-over-Year Comparison (avg. across selected districts)</p>',unsafe_allow_html=True)
        years_in_df = sorted(df['Year'].unique().tolist())
        if len(years_in_df) < 2:
            st.info("Select at least two years in the Year filter to see comparison.")
        else:
            # Chunk summary cards into rows of at most 4 so they never get squeezed
            CARD_CHUNK = 4
            for start in range(0, len(years_in_df), CARD_CHUNK):
                chunk = years_in_df[start:start+CARD_CHUNK]
                ycols = st.columns(len(chunk))
                for i, yr in enumerate(chunk):
                    ydf = df[df['Year']==yr]
                    with ycols[i]:
                        h_n  = int((ydf['Predicted_Class']=='High').sum())
                        l_n  = int((ydf['Predicted_Class']=='Low').sum())
                        ar   = ydf['RiskScore'].mean()
                        rain = ydf['Rainfall'].mean()
                        st.markdown(f"""
                        <div class="section-card" style="border-top:4px solid #2B6CB0;">
                            <div class="section-title">📅 {yr}</div>
                            <div class="lrow"><span class="lkey">High Risk Records</span><span class="lval" style="color:#B84A20">{h_n}</span></div>
                            <div class="lrow"><span class="lkey">Low Risk Records</span><span class="lval" style="color:#2A7A3B">{l_n}</span></div>
                            <div class="lrow"><span class="lkey">Avg Risk Score</span><span class="lval">{ar:.1f}</span></div>
                            <div class="lrow"><span class="lkey">Avg Rainfall</span><span class="lval">{rain:.1f} mm</span></div>
                        </div>""", unsafe_allow_html=True)
                st.markdown("<div style='margin-bottom:10px;'></div>", unsafe_allow_html=True)

        # aggregate across districts per Month/Year before plotting
        d2 = df.groupby(['Year','Month','Month_Num'], as_index=False).agg(
            RiskScore=('RiskScore','mean'), Rainfall=('Rainfall','mean')
        ).sort_values(['Year','Month_Num'])

        # Line charts (one line per year) instead of grouped bars — far less
        # cramped once more than 2 years are selected, and easier to read trends from.
        palette = px.colors.qualitative.Set2

        fig_yc = go.Figure()
        for i, yr in enumerate(sorted(d2['Year'].unique())):
            yd = d2[d2['Year']==yr].sort_values('Month_Num')
            fig_yc.add_trace(go.Scatter(
                x=yd['Month'], y=yd['RiskScore'], mode='lines+markers',
                name=str(yr), line=dict(width=3, color=palette[i % len(palette)]),
                marker=dict(size=8, line=dict(color='white', width=1)),
                hovertemplate=f'<b>%{{x}} {yr}</b><br>Avg Risk Score: %{{y:.1f}}<extra></extra>',
            ))
        fig_yc.update_layout(**_layout(
            title='Monthly Avg. Risk Score — Year Comparison',
            yaxis_title='Avg Risk Score', xaxis_tickangle=-35, height=440,
            xaxis=dict(categoryorder='array', categoryarray=MONTH_ORDER),
            legend=dict(orientation='h', y=-0.22, x=0.5, xanchor='center'),
        ))
        st.plotly_chart(fig_yc, use_container_width=True, config={'displayModeBar':False})

        fig_rain = go.Figure()
        for i, yr in enumerate(sorted(d2['Year'].unique())):
            yd = d2[d2['Year']==yr].sort_values('Month_Num')
            fig_rain.add_trace(go.Scatter(
                x=yd['Month'], y=yd['Rainfall'], mode='lines+markers',
                name=str(yr), line=dict(width=3, color=palette[i % len(palette)], dash='dot'),
                marker=dict(size=8, line=dict(color='white', width=1)),
                hovertemplate=f'<b>%{{x}} {yr}</b><br>Avg Rainfall: %{{y:.1f}} mm<extra></extra>',
            ))
        fig_rain.update_layout(**_layout(
            title='Monthly Avg. Rainfall — Year Comparison',
            yaxis_title='Rainfall (mm)', xaxis_tickangle=-35, height=440,
            xaxis=dict(categoryorder='array', categoryarray=MONTH_ORDER),
            legend=dict(orientation='h', y=-0.22, x=0.5, xanchor='center'),
        ))
        st.plotly_chart(fig_rain, use_container_width=True, config={'displayModeBar':False})
        st.markdown('</div>',unsafe_allow_html=True)

    # ════════ TAB 5 — Indicator Deep Dive ════════
    if active_tab == TAB_LABELS[4]:
        st.markdown('<div class="section-card">',unsafe_allow_html=True)
        st.markdown('<p class="section-title">🌡️ Indicator Deep Dive — Monthly Trends (avg. across selected districts)</p>',unsafe_allow_html=True)

        indicators = {
            'Rainfall':      ('Rainfall (mm)',    '#3182CE', 0,    500),
            'Temperature':   ('Temperature (°C)', '#E53E3E', 20,   45),
            'NDVI':          ('NDVI Index',        '#38A169', 0,    1),
            'Soil_Moisture': ('Soil Moisture',     '#744210', 0,    1),
            'SPI3':          ('SPI-3 Index',       '#2B6CB0', -3,   3),
        }
        d3 = df.groupby(['Year','Month','Month_Num'], as_index=False)[
            [c for c in indicators if c in df.columns]
        ].mean()

        for col, (label, color, ymin, ymax) in indicators.items():
            if col not in df.columns:
                continue
            fig_ind = go.Figure()
            for yr in sorted(d3['Year'].unique()):
                yd = d3[d3['Year']==yr].sort_values('Month_Num')
                fig_ind.add_trace(go.Scatter(
                    x=yd['Month'], y=yd[col],
                    mode='lines+markers', name=str(yr),
                    line=dict(width=2.5),
                    marker=dict(size=8),
                    hovertemplate=f'<b>%{{x}} {yr}</b><br>{label}: %{{y:.3f}}<extra></extra>',
                ))
            if col == 'SPI3':
                fig_ind.add_hline(y=-1, line_dash='dash', line_color='#B84A20',
                                  annotation_text='Drought threshold',
                                  annotation_font=dict(size=9, color='#B84A20'))
            if col == 'Rainfall':
                fig_ind.add_hline(y=50, line_dash='dash', line_color='#C47B1A',
                                  annotation_text='Low rainfall warning',
                                  annotation_font=dict(size=9, color='#C47B1A'))
            fig_ind.update_layout(**_layout(
                title=f'{label} — Monthly Trend',
                yaxis_title=label,
                yaxis_range=[ymin, ymax],
                xaxis=dict(categoryorder='array', categoryarray=MONTH_ORDER),
            ))
            st.plotly_chart(fig_ind, use_container_width=True, config={'displayModeBar':False})

        st.markdown('</div>',unsafe_allow_html=True)

    # ════════ TAB 4 — Seasonal Analysis ════════
    if active_tab == TAB_LABELS[3]:
        st.markdown('<div class="section-card">',unsafe_allow_html=True)
        st.markdown('<p class="section-title">📅 Seasonal Analysis — Kharif · Rabi · Zaid</p>',unsafe_allow_html=True)

        season_map = {
            'JUN':'Kharif','JUL':'Kharif','AUG':'Kharif','SEP':'Kharif','OCT':'Kharif',
            'NOV':'Rabi','DEC':'Rabi','JAN':'Rabi','FEB':'Rabi',
            'MAR':'Zaid','APR':'Zaid','MAY':'Zaid',
        }
        season_colors = {'Kharif':'#38A169','Rabi':'#C47B1A','Zaid':'#E53E3E'}

        d4 = df.copy()
        d4['Season'] = d4['Month'].map(season_map).fillna('Other')

        sc1,sc2,sc3 = st.columns(3)
        for col_s, season in zip([sc1,sc2,sc3],['Kharif','Rabi','Zaid']):
            sdf = d4[d4['Season']==season]
            if sdf.empty:
                continue
            h_pct  = (sdf['Predicted_Class']=='High').mean()*100
            ar     = sdf['RiskScore'].mean()
            rain   = sdf['Rainfall'].mean()
            sc_clr = season_colors[season]
            with col_s:
                st.markdown(f"""
                <div class="section-card" style="border-top:4px solid {sc_clr};text-align:center;">
                    <div class="section-title">{season}</div>
                    <div class="kpi-value" style="color:{sc_clr}">{ar:.1f}</div>
                    <div class="kpi-sub">Avg Risk Score</div>
                    <br>
                    <div class="lrow"><span class="lkey">High Risk %</span><span class="lval">{h_pct:.0f}%</span></div>
                    <div class="lrow"><span class="lkey">Avg Rainfall</span><span class="lval">{rain:.1f} mm</span></div>
                    <div class="lrow"><span class="lkey">Records</span><span class="lval">{len(sdf)}</span></div>
                </div>""", unsafe_allow_html=True)

        season_agg = d4.groupby(['Season','Year']).agg(
            AvgRisk=('RiskScore','mean'),
            AvgRainfall=('Rainfall','mean'),
            HighRiskCount=('Predicted_Class', lambda x: (x=='High').sum())
        ).reset_index().round(1)

        fig_s1 = px.bar(season_agg, x='Season', y='AvgRisk', color='Year',
                        barmode='group', text='AvgRisk',
                        color_discrete_sequence=['#2B6CB0','#C47B1A'],
                        category_orders={'Season':['Kharif','Rabi','Zaid']})
        fig_s1.update_traces(texttemplate='%{text:.1f}', textposition='inside',
                             textfont=dict(color='white', size=11))
        fig_s1.update_layout(**_layout(title='Average Risk Score by Season',
                                        yaxis_title='Avg Risk Score'))
        st.plotly_chart(fig_s1, use_container_width=True, config={'displayModeBar':False})

        fig_s2 = px.bar(season_agg, x='Season', y='AvgRainfall', color='Year',
                        barmode='group', text='AvgRainfall',
                        color_discrete_sequence=['#3182CE','#68D391'],
                        category_orders={'Season':['Kharif','Rabi','Zaid']})
        fig_s2.update_traces(texttemplate='%{text:.0f}', textposition='inside',
                             textfont=dict(color='white', size=11))
        fig_s2.update_layout(**_layout(title='Average Rainfall by Season (mm)',
                                        yaxis_title='Rainfall (mm)'))
        st.plotly_chart(fig_s2, use_container_width=True, config={'displayModeBar':False})
        st.markdown('</div>',unsafe_allow_html=True)

    # ════════ TAB 6 — Month Rankings ════════
    if active_tab == TAB_LABELS[5]:
        st.markdown('<div class="section-card">',unsafe_allow_html=True)
        st.markdown('<p class="section-title">🏆 District-Month Rankings — Severity Order</p>',unsafe_allow_html=True)

        d5 = df[['District','Year','Month','Predicted_Class','RiskScore','RiskLevel',
                  'Confidence','Rainfall','SPI3']].copy()
        d5['Period'] = d5['District'] + ' · ' + d5['Month'] + ' ' + d5['Year'].astype(str)
        d5 = d5.sort_values('RiskScore', ascending=False).reset_index(drop=True)
        d5.index += 1

        st.markdown("#### 🔴 Top 5 Most Severe Drought District-Months")
        top5 = d5.head(5)
        for rank, row in top5.iterrows():
            clr = '#B84A20' if row['RiskLevel']=='High' else '#C47B1A'
            st.markdown(f"""
            <div style="background:#FFF5F5;border-left:5px solid {clr};
                        border-radius:8px;padding:10px 16px;margin-bottom:8px;
                        display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;">
                <span style="font-size:1.4rem;font-weight:800;color:{clr};">#{rank}</span>
                <span style="font-weight:700;color:#1A202C;font-size:0.95rem;">{row['Period']}</span>
                <span style="color:#4A5568;">Class: <b>{row['Predicted_Class']}</b></span>
                <span style="color:#4A5568;">Risk Score: <b style="color:{clr}">{row['RiskScore']:.1f}</b></span>
                <span style="color:#4A5568;">Rainfall: <b>{row['Rainfall']:.1f} mm</b></span>
                <span style="color:#4A5568;">SPI-3: <b>{row['SPI3']:.3f}</b></span>
            </div>""", unsafe_allow_html=True)

        st.markdown("#### 🟢 Top 5 Best (Lowest Risk) District-Months")
        bot5 = d5.tail(5).iloc[::-1]
        for rank, row in bot5.iterrows():
            st.markdown(f"""
            <div style="background:#F0FFF4;border-left:5px solid #38A169;
                        border-radius:8px;padding:10px 16px;margin-bottom:8px;
                        display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;">
                <span style="font-size:1.4rem;font-weight:800;color:#38A169;">#{rank}</span>
                <span style="font-weight:700;color:#1A202C;font-size:0.95rem;">{row['Period']}</span>
                <span style="color:#4A5568;">Class: <b>{row['Predicted_Class']}</b></span>
                <span style="color:#4A5568;">Risk Score: <b style="color:#38A169">{row['RiskScore']:.1f}</b></span>
                <span style="color:#4A5568;">Rainfall: <b>{row['Rainfall']:.1f} mm</b></span>
                <span style="color:#4A5568;">SPI-3: <b>{row['SPI3']:.3f}</b></span>
            </div>""", unsafe_allow_html=True)

        st.markdown("#### 📋 Full Rankings Table")
        d5_show = d5.copy()
        d5_show['Confidence'] = (d5_show['Confidence']*100).round(1).astype(str)+'%'
        d5_show['RiskScore']  = d5_show['RiskScore'].round(1)
        d5_show['Rainfall']   = d5_show['Rainfall'].round(1)
        d5_show['SPI3']       = d5_show['SPI3'].round(3)
        d5_show = d5_show[['Period','District','Predicted_Class','RiskScore','RiskLevel',
                             'Confidence','Rainfall','SPI3']]
        d5_show.columns = ['Period','District','Predicted Class','Risk Score','Risk Level',
                            'Confidence','Rainfall (mm)','SPI-3']
        st.dataframe(d5_show, use_container_width=True, height=400, hide_index=False)
        st.markdown('</div>',unsafe_allow_html=True)

    # ════════ TAB 9 — Advisory Board ════════
    if active_tab == TAB_LABELS[8]:
        st.markdown('<div class="section-card">',unsafe_allow_html=True)
        st.markdown('<p class="section-title">📋 Advisory Board — Action Required District-Months</p>',unsafe_allow_html=True)

        if 'Advisory' not in df.columns:
            st.info("No Advisory column found in your CSV.")
        else:
            d6 = df.copy()
            d6['Period'] = d6['District'] + ' · ' + d6['Month'] + ' ' + d6['Year'].astype(str)
            d6 = d6.sort_values(['Year','Month_Num'])

            a1,a2,a3 = st.columns(3)
            high_df = d6[d6['RiskLevel']=='High']
            mod_df  = d6[d6['RiskLevel']=='Moderate']
            low_df  = d6[d6['RiskLevel']=='Low']
            with a1:
                st.markdown(f"""<div class="kpi-card rust">
                    <div class="kpi-label">🔴 High Risk Records</div>
                    <div class="kpi-value">{len(high_df)}</div>
                    <div class="kpi-sub">Immediate action needed</div>
                </div>""", unsafe_allow_html=True)
            with a2:
                st.markdown(f"""<div class="kpi-card amber">
                    <div class="kpi-label">🟠 Moderate Risk Records</div>
                    <div class="kpi-value">{len(mod_df)}</div>
                    <div class="kpi-sub">Monitor closely</div>
                </div>""", unsafe_allow_html=True)
            with a3:
                st.markdown(f"""<div class="kpi-card green">
                    <div class="kpi-label">🟢 Low Risk Records</div>
                    <div class="kpi-value">{len(low_df)}</div>
                    <div class="kpi-sub">Normal conditions</div>
                </div>""", unsafe_allow_html=True)

            st.markdown("<br>",unsafe_allow_html=True)

            MAX_CARDS = 60  # keep the tab responsive when many districts/months are selected

            if not high_df.empty:
                st.markdown(f"### 🔴 High Risk — Immediate Action Required (showing up to {MAX_CARDS})")
                for _, row in high_df.head(MAX_CARDS).iterrows():
                    advisory = str(row['Advisory']).replace('\\n',' ').strip()
                    st.markdown(f"""
                    <div style="background:#FFF5F5;border:1px solid #FED7D7;
                                border-left:5px solid #B84A20;border-radius:8px;
                                padding:14px 18px;margin-bottom:10px;">
                        <div style="display:flex;justify-content:space-between;margin-bottom:6px;flex-wrap:wrap;">
                            <span style="font-weight:700;color:#B84A20;font-size:1.0rem;">
                                📍 {row['Period']}
                            </span>
                            <span style="background:#B84A20;color:#fff;border-radius:20px;
                                         padding:2px 12px;font-size:0.72rem;font-weight:600;">
                                Risk Score: {row['RiskScore']:.1f}
                            </span>
                        </div>
                        <div style="color:#744210;font-size:0.82rem;margin-bottom:6px;">
                            <b>Class:</b> {row['Predicted_Class']} &nbsp;·&nbsp;
                            <b>Confidence:</b> {row['Confidence']*100:.1f}% &nbsp;·&nbsp;
                            <b>Rainfall:</b> {row['Rainfall']:.1f} mm &nbsp;·&nbsp;
                            <b>SPI-3:</b> {row['SPI3']:.3f}
                        </div>
                        <div style="color:#1A202C;font-size:0.84rem;line-height:1.55;">
                            📌 {advisory}
                        </div>
                    </div>""", unsafe_allow_html=True)

            if not mod_df.empty:
                st.markdown(f"### 🟠 Moderate Risk — Monitor Closely (showing up to {MAX_CARDS})")
                for _, row in mod_df.head(MAX_CARDS).iterrows():
                    advisory = str(row['Advisory']).replace('\\n',' ').strip()
                    st.markdown(f"""
                    <div style="background:#FFFBEB;border:1px solid #FEF3C7;
                                border-left:5px solid #C47B1A;border-radius:8px;
                                padding:14px 18px;margin-bottom:10px;">
                        <div style="display:flex;justify-content:space-between;margin-bottom:6px;flex-wrap:wrap;">
                            <span style="font-weight:700;color:#C47B1A;font-size:1.0rem;">
                                📍 {row['Period']}
                            </span>
                            <span style="background:#C47B1A;color:#fff;border-radius:20px;
                                         padding:2px 12px;font-size:0.72rem;font-weight:600;">
                                Risk Score: {row['RiskScore']:.1f}
                            </span>
                        </div>
                        <div style="color:#744210;font-size:0.82rem;margin-bottom:6px;">
                            <b>Class:</b> {row['Predicted_Class']} &nbsp;·&nbsp;
                            <b>Confidence:</b> {row['Confidence']*100:.1f}% &nbsp;·&nbsp;
                            <b>Rainfall:</b> {row['Rainfall']:.1f} mm &nbsp;·&nbsp;
                            <b>SPI-3:</b> {row['SPI3']:.3f}
                        </div>
                        <div style="color:#1A202C;font-size:0.84rem;line-height:1.55;">
                            📌 {advisory}
                        </div>
                    </div>""", unsafe_allow_html=True)

            if not low_df.empty:
                st.markdown(f"### 🟢 Low Risk — Normal Conditions (showing up to {MAX_CARDS})")
                for _, row in low_df.head(MAX_CARDS).iterrows():
                    advisory = str(row['Advisory']).replace('\\n',' ').strip()
                    st.markdown(f"""
                    <div style="background:#F0FFF4;border:1px solid #C6F6D5;
                                border-left:5px solid #38A169;border-radius:8px;
                                padding:14px 18px;margin-bottom:10px;">
                        <div style="display:flex;justify-content:space-between;margin-bottom:6px;flex-wrap:wrap;">
                            <span style="font-weight:700;color:#38A169;font-size:1.0rem;">
                                📍 {row['Period']}
                            </span>
                            <span style="background:#38A169;color:#fff;border-radius:20px;
                                         padding:2px 12px;font-size:0.72rem;font-weight:600;">
                                Risk Score: {row['RiskScore']:.1f}
                            </span>
                        </div>
                        <div style="color:#276749;font-size:0.82rem;margin-bottom:6px;">
                            <b>Class:</b> {row['Predicted_Class']} &nbsp;·&nbsp;
                            <b>Confidence:</b> {row['Confidence']*100:.1f}% &nbsp;·&nbsp;
                            <b>Rainfall:</b> {row['Rainfall']:.1f} mm &nbsp;·&nbsp;
                            <b>SPI-3:</b> {row['SPI3']:.3f}
                        </div>
                        <div style="color:#1A202C;font-size:0.84rem;line-height:1.55;">
                            📌 {advisory}
                        </div>
                    </div>""", unsafe_allow_html=True)

        st.markdown('</div>',unsafe_allow_html=True)

    # ════════ TAB 10 — Live District Model Prediction ════════
    if active_tab == TAB_LABELS[9]:
        model     = load_district_model(DISTRICT_MODEL_PATH)
        d_features = load_district_features(DISTRICT_FEATURES_PATH)

        if model is None:
            st.markdown('<div class="section-card">', unsafe_allow_html=True)
            st.markdown('<p class="section-title">🤖 Live District Model Prediction</p>', unsafe_allow_html=True)
            st.error(
                f"⚠️ Model file not found at `{DISTRICT_MODEL_PATH}`.\n\n"
                "Place your trained `Telangana_District_Drought_Model.pkl` (output of "
                "`train_drought_model.py`) in the same folder as this script."
            )
            st.info(f"Expected feature columns (in order): `{', '.join(d_features)}`")
            st.markdown('</div>', unsafe_allow_html=True)
        else:
            pred_tab1, pred_tab2 = st.tabs(["✍️ Manual Entry", "📁 Upload CSV"])

            # ── Manual Entry ──
            with pred_tab1:
                st.markdown('<div class="section-card">', unsafe_allow_html=True)
                st.markdown('<p class="section-title">✍️ Manual Feature Entry</p>', unsafe_allow_html=True)
                st.caption(
                    "Enter this month's indicators for a district. Previous-month values are "
                    "pulled automatically from that district's most recent record, and the "
                    "Groundwater Proxy is estimated automatically from your inputs — you don't "
                    "need to enter either by hand."
                )

                ref_district = st.selectbox("District", options=districts_all, key="manual_ref_district")

                gw_model_obj_preview = load_gw_model()
                LAG_MAP = {'Rainfall_lag1':'Rainfall', 'SoilMoisture_lag1':'Soil_Moisture', 'NDVI_lag1':'NDVI'}
                base_features = [f for f in d_features if f in ('Rainfall','Temperature','Soil_Moisture','NDVI')]
                lag_features  = [f for f in d_features if f in LAG_MAP]
                has_gw_feature = 'Groundwater_Proxy' in d_features
                # Only ask the user for Groundwater Proxy manually if we have no GW model to estimate it
                show_manual_gw = has_gw_feature and (gw_model_obj_preview is None)

                entry_features = base_features + (['Groundwater_Proxy'] if show_manual_gw else [])
                defaults = {"Rainfall": 50.0, "Temperature": 32.0, "Soil_Moisture": 0.25,
                            "NDVI": 0.40, "Groundwater_Proxy": 0.50}

                # ── Optional: fetch this district's current-month indicators from GEE ──
                if GEE_LIVE_AVAILABLE:
                    fetch_col, status_col = st.columns([1, 3])
                    with fetch_col:
                        fetch_clicked = st.button("🌍 Fetch Live from Earth Engine",
                                                   use_container_width=True,
                                                   key="gee_live_fetch_btn")
                    if fetch_clicked:
                        with st.spinner(f"Pulling latest satellite data for {ref_district}…"):
                            try:
                                live_result = gee_live_fetch.fetch_live_indicators(
                                    geojson_path=DISTRICT_GEOJSON_PATH,
                                    district_name=ref_district,
                                )
                                st.session_state["gee_live_values"] = live_result
                                st.session_state["gee_live_district"] = ref_district
                            except Exception as e:
                                st.session_state.pop("gee_live_values", None)
                                with status_col:
                                    st.error(f"Live fetch failed: {e}")

                    live_vals = st.session_state.get("gee_live_values", {})
                    live_for_this_district = (
                        live_vals if st.session_state.get("gee_live_district") == ref_district else {}
                    )
                    if live_for_this_district:
                        with status_col:
                            st.success(
                                f"✅ Loaded live indicators for **{ref_district}** — "
                                f"{live_for_this_district.get('Period','')} "
                                "(satellite data typically lags 1–2 months behind real time). "
                                "Values below are pre-filled — adjust if needed before running."
                            )
                else:
                    live_for_this_district = {}
                    st.caption(
                        "🌍 Live Earth Engine fetch not available in this environment "
                        "(gee_live_fetch module not found or Earth Engine not configured). "
                        "Enter indicators manually below."
                    )

                with st.form("manual_predict_form"):
                    cols = st.columns(3)
                    manual_vals = {}
                    for i, feat in enumerate(entry_features):
                        with cols[i % 3]:
                            prefill = live_for_this_district.get(feat, defaults.get(feat, 0.0))
                            manual_vals[feat] = st.number_input(
                                FEATURE_LABELS.get(feat, feat),
                                value=prefill,
                                format="%.3f",
                            )
                    run_btn = st.form_submit_button("🔮 Run Prediction", use_container_width=True)

                if run_btn:
                    try:
                        auto_notes = {}

                        # ── Auto-derive lag features from this district's latest historical record ──
                        hist = df_raw[df_raw['District']==ref_district].sort_values(['Year','Month_Num'])
                        last_row = hist.iloc[-1] if not hist.empty else None
                        for lf in lag_features:
                            base_col = LAG_MAP[lf]
                            if last_row is not None and base_col in last_row.index and pd.notna(last_row[base_col]):
                                manual_vals[lf] = float(last_row[base_col])
                                auto_notes[lf] = f"{manual_vals[lf]:.3f} — from {last_row['Period']} record"
                            else:
                                manual_vals[lf] = manual_vals.get(base_col, defaults.get(base_col, 0.0))
                                auto_notes[lf] = f"{manual_vals[lf]:.3f} — no history found, assumed same as current month"

                        # ── Auto-derive Groundwater Proxy via the GW model ──
                        gw_model_obj = load_gw_model()
                        if has_gw_feature and not show_manual_gw and gw_model_obj is not None:
                            gw_feat_row = {
                                "Rainfall":     manual_vals["Rainfall"],
                                "Temperature":  manual_vals["Temperature"],
                                "SoilMoisture": manual_vals["Soil_Moisture"],
                                "NDVI":         manual_vals["NDVI"],
                            }
                            gw_val, gw_status_auto, _ = run_gw_prediction(gw_model_obj, gw_feat_row)
                            manual_vals["Groundwater_Proxy"] = gw_val
                            auto_notes["Groundwater_Proxy"] = f"{gw_val:.3f} — auto-estimated ({gw_status_auto})"

                        pred, conf, risk_s = run_prediction(model, manual_vals, d_features)
                        risk_l = 'High' if risk_s>=70 else ('Moderate' if risk_s>=40 else 'Low')
                        cls_color  = CLASS_COLORS.get(str(pred), '#1A365D')
                        risk_color = RISK_COLORS.get(risk_l, '#1A365D')
                        conf_txt   = f"{conf*100:.1f}%" if conf is not None else "N/A"

                        gw_final = manual_vals.get("Groundwater_Proxy")
                        if gw_final is not None:
                            if gw_final >= 0.66:
                                gw_status, gw_color = "Adequate", "#38A169"
                            elif gw_final >= 0.33:
                                gw_status, gw_color = "Moderate", "#DD6B20"
                            else:
                                gw_status, gw_color = "Critically Low", "#E53E3E"
                        else:
                            gw_status, gw_color = "N/A", "#718096"

                        col_d, col_g = st.columns(2)
                        with col_d:
                            st.markdown(f"""
                            <div class="lookup-box">
                                <h4>🔮 Drought Prediction — {ref_district}</h4>
                                <div class="lrow"><span class="lkey">Predicted Class</span>
                                    <span class="lval" style="color:{cls_color}">● {pred}</span></div>
                                <div class="lrow"><span class="lkey">Model Confidence</span>
                                    <span class="lval">{conf_txt}</span></div>
                                <div class="lrow"><span class="lkey">Estimated Risk Score</span>
                                    <span class="lval">{risk_s:.1f} / 100</span></div>
                                <div class="lrow"><span class="lkey">Risk Level</span>
                                    <span class="lval" style="color:{risk_color}">▲ {risk_l}</span></div>
                            </div>""", unsafe_allow_html=True)
                        with col_g:
                            gw_display = f"{gw_final:.4f}" if gw_final is not None else "N/A"
                            gw_model_label = "RandomForest Regressor (auto-estimated)" if (has_gw_feature and not show_manual_gw and gw_model_obj is not None) else "Manually entered"
                            st.markdown(f"""
                            <div class="lookup-box">
                                <h4>💧 Groundwater Level</h4>
                                <div class="lrow"><span class="lkey">Groundwater Proxy</span>
                                    <span class="lval">{gw_display}</span></div>
                                <div class="lrow"><span class="lkey">Status</span>
                                    <span class="lval" style="color:{gw_color}">● {gw_status}</span></div>
                                <div class="lrow"><span class="lkey">Scale</span>
                                    <span class="lval">0.00 (depleted) → 1.00 (adequate)</span></div>
                                <div class="lrow"><span class="lkey">Source</span>
                                    <span class="lval">{gw_model_label}</span></div>
                            </div>""", unsafe_allow_html=True)

                        if auto_notes:
                            with st.expander("ℹ️ Auto-filled values used in this prediction"):
                                for k, v in auto_notes.items():
                                    st.markdown(f"- **{FEATURE_LABELS.get(k,k)}**: {v}")
                    except Exception as e:
                        st.error(f"Prediction failed: {e}")
                        st.caption("Check that District_Model_Features.pkl matches your model's expected input columns exactly.")
                st.markdown('</div>', unsafe_allow_html=True)

            # ── CSV Upload ──
            with pred_tab2:
                st.markdown('<div class="section-card">', unsafe_allow_html=True)
                st.markdown('<p class="section-title">📁 Batch Prediction from CSV</p>', unsafe_allow_html=True)
                st.caption(
                    f"Upload a CSV containing the district model columns: `{', '.join(d_features)}` "
                    f"(a `District` column is optional but recommended for labeling) — "
                    f"the Groundwater model also runs automatically if Rainfall, Temperature, "
                    f"Soil_Moisture and NDVI are present."
                )

                uploaded = st.file_uploader("Choose CSV file", type=["csv"], key="batch_upload_district")

                if uploaded is not None:
                    try:
                        up_df = pd.read_csv(uploaded)
                        missing = [c for c in d_features if c not in up_df.columns]
                        if missing:
                            st.error(f"Missing required columns for district model: {', '.join(missing)}")
                        else:
                            X_batch = up_df[d_features]
                            raw_preds = model.predict(X_batch)
                            preds = [_label_for(p) for p in raw_preds]
                            if hasattr(model, "predict_proba"):
                                probas = model.predict_proba(X_batch)
                                confs  = probas.max(axis=1)
                            else:
                                confs = [None]*len(preds)

                            risk_map = {"High":1.0,"Moderate":0.5,"Low":0.15}
                            result_df = up_df.copy()
                            result_df['Predicted_Class'] = preds
                            result_df['Confidence']      = [round(float(c),4) if c is not None else None for c in confs]
                            result_df['RiskScore']       = [
                                round(risk_map.get(str(p),0.5) * (c if c else 0.7) * 100, 1)
                                for p, c in zip(preds, confs)
                            ]
                            result_df['RiskLevel'] = result_df['RiskScore'].apply(
                                lambda s: 'High' if s>=70 else ('Moderate' if s>=40 else 'Low'))

                            gw_model_obj = load_gw_model()
                            gw_input_df  = up_df.rename(columns={'Soil_Moisture':'SoilMoisture'})
                            gw_missing   = [c for c in GW_MODEL_FEATURES if c not in gw_input_df.columns]
                            if gw_model_obj is not None and not gw_missing:
                                X_gw = gw_input_df[GW_MODEL_FEATURES]
                                gw_preds = gw_model_obj.predict(X_gw)
                                gw_preds = np.clip(gw_preds, 0.0, 1.0).round(4)
                                result_df['Groundwater_Proxy_Predicted'] = gw_preds
                                result_df['Groundwater_Status'] = result_df['Groundwater_Proxy_Predicted'].apply(
                                    lambda v: 'Adequate' if v>=0.66 else ('Moderate' if v>=0.33 else 'Critically Low')
                                )
                                gw_note = "✅ Groundwater predictions added."
                            elif gw_model_obj is None:
                                gw_note = "⚠️ Groundwater_Model.pkl not found — skipped."
                            else:
                                gw_note = f"⚠️ Groundwater model skipped — missing columns: {', '.join(gw_missing)}"

                            st.success(f"✅ {len(result_df)} district predictions generated.  {gw_note}")
                            st.dataframe(result_df, use_container_width=True, height=320, hide_index=True)

                            out_buf = BytesIO()
                            with pd.ExcelWriter(out_buf, engine='openpyxl') as writer:
                                result_df.to_excel(writer, index=False, sheet_name='Predictions')
                            st.download_button(
                                "⬇️ Download Predictions (Excel)",
                                data=out_buf.getvalue(),
                                file_name=f"Batch_District_Predictions_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True,
                            )
                    except Exception as e:
                        st.error(f"Could not process file: {e}")
                st.markdown('</div>', unsafe_allow_html=True)

    # ════════ TAB 11 — Why This Prediction? (SHAP) ════════
    if active_tab == TAB_LABELS[10]:
        st.markdown('<div class="section-card">', unsafe_allow_html=True)
        st.markdown('<p class="section-title">🧬 Why Did the Model Predict This? (SHAP Explainability)</p>',
                    unsafe_allow_html=True)
        st.caption(
            "SHAP (SHapley Additive exPlanations) shows exactly which features pushed the "
            "district model toward or away from its prediction, and by how much. "
            "🔴 Red bars push toward the predicted class · 🟢 Green bars push away from it."
        )

        model      = load_district_model(DISTRICT_MODEL_PATH)
        d_features = load_district_features(DISTRICT_FEATURES_PATH)

        if model is None:
            st.error(f"⚠️ Model not found at `{DISTRICT_MODEL_PATH}`. Load your `.pkl` to use this tab.")
            st.markdown('</div>', unsafe_allow_html=True)
        else:
            try:
                import shap  # noqa: F401
                shap_available = True
            except ImportError:
                shap_available = False

            if not shap_available:
                st.warning(
                    "The `shap` library is not installed. Run this in your terminal:\n\n"
                    "```\npip install shap\n```\n\nThen restart the dashboard."
                )
                st.markdown('</div>', unsafe_allow_html=True)
            else:
                explain_year     = st.selectbox("Select Year",     options=years_all,     key="shap_year")
                explain_month    = st.selectbox("Select Month",    options=months_all,    key="shap_month")
                explain_district = st.selectbox("Select District", options=districts_all, key="shap_district")

                row_match = df_raw[(df_raw['Year']==explain_year) &
                                    (df_raw['Month']==explain_month) &
                                    (df_raw['District']==explain_district)]
                missing_feats = [f for f in d_features if f not in df_raw.columns]

                if missing_feats:
                    st.error(f"Dataset is missing required model columns: {', '.join(missing_feats)}")
                elif row_match.empty:
                    st.info(f"No data found for {explain_district} — {explain_month} {explain_year}.")
                else:
                    feature_row = {f: float(row_match.iloc[0][f]) for f in d_features}
                    actual_risk  = row_match.iloc[0]['RiskScore']
                    actual_level = row_match.iloc[0]['RiskLevel']

                    if st.button("🔍 Explain This Prediction", use_container_width=True):
                        with st.spinner("Computing SHAP values… this may take a few seconds."):
                            try:
                                explainer = build_shap_explainer(model, df_raw, d_features)
                                shap_df, pred_class = explain_prediction(explainer, model, feature_row, d_features)

                                X_one = pd.DataFrame([feature_row])[d_features]
                                conf_val = None
                                if hasattr(model, "predict_proba"):
                                    proba = model.predict_proba(X_one)[0]
                                    conf_val = float(np.max(proba)) * 100

                                cls_color  = CLASS_COLORS.get(str(pred_class), '#1A365D')
                                risk_color = RISK_COLORS.get(actual_level, '#1A365D')

                                st.markdown(f"""
                                <div class="lookup-box">
                                    <h4>📍 {explain_district} — {MONTH_FULL.get(explain_month, explain_month)} {explain_year} — Prediction Summary</h4>
                                    <div class="lrow"><span class="lkey">Class</span>
                                        <span class="lval" style="color:{cls_color};font-size:1.05rem;">● {pred_class}</span></div>
                                    <div class="lrow"><span class="lkey">Confidence</span>
                                        <span class="lval">{f'{conf_val:.0f}%' if conf_val is not None else 'N/A'}</span></div>
                                    <div class="lrow"><span class="lkey">Risk Score</span>
                                        <span class="lval">{actual_risk:.0f}</span></div>
                                    <div class="lrow"><span class="lkey">Risk Level</span>
                                        <span class="lval" style="color:{risk_color}">▲ {actual_level}</span></div>
                                </div>""", unsafe_allow_html=True)

                                st.markdown("<br>", unsafe_allow_html=True)

                                total_abs = shap_df['AbsImpact'].sum()
                                shap_df['PctContribution'] = (
                                    (shap_df['AbsImpact'] / total_abs * 100) if total_abs > 0 else 0
                                )

                                st.markdown("#### 🧠 Why did the AI predict this?")
                                checklist_html = '<div class="lookup-box">'
                                for _, r in shap_df.iterrows():
                                    direction_color = '#E53E3E' if r['SHAP_Value'] > 0 else '#38A169'
                                    checklist_html += f"""
                                    <div class="lrow">
                                        <span class="lkey">✔ {r['Feature']}</span>
                                        <span class="lval" style="color:{direction_color}">
                                            contributed {r['PctContribution']:.0f}%
                                        </span>
                                    </div>"""
                                checklist_html += '</div>'
                                st.markdown(checklist_html, unsafe_allow_html=True)

                                st.markdown("<br>", unsafe_allow_html=True)

                                st.markdown("#### 🌊 Waterfall Plot")
                                fig_shap = chart_shap_waterfall(shap_df, 0.0, str(pred_class))
                                st.plotly_chart(fig_shap, use_container_width=True,
                                                config={'displayModeBar': False})

                                st.markdown("#### 📊 Contribution Bar Chart")
                                bar_df = shap_df.sort_values('PctContribution', ascending=True)
                                bar_colors = bar_df['SHAP_Value'].apply(
                                    lambda v: '#E53E3E' if v > 0 else '#38A169')
                                fig_pct = go.Figure(go.Bar(
                                    x=bar_df['PctContribution'], y=bar_df['Feature'],
                                    orientation='h', marker_color=bar_colors, marker_line_width=0,
                                    text=bar_df['PctContribution'].apply(lambda v: f"{v:.0f}%"),
                                    textposition='inside', textfont=dict(color='white', size=12),
                                    hovertemplate='<b>%{y}</b><br>Contribution: %{x:.1f}%<extra></extra>',
                                ))
                                fig_pct.update_layout(**_layout(
                                    title='Relative Feature Contribution (%)',
                                    xaxis_title='Contribution to Prediction (%)',
                                    height=340, margin=dict(l=200, r=40, t=42, b=12),
                                ))
                                st.plotly_chart(fig_pct, use_container_width=True,
                                                config={'displayModeBar': False})

                                st.markdown("#### 📝 Plain-Language Summary")
                                top_pos = shap_df[shap_df['SHAP_Value'] > 0].head(3)
                                top_neg = shap_df[shap_df['SHAP_Value'] < 0].head(3)

                                if not top_pos.empty:
                                    reasons = ", ".join(
                                        f"**{r['Feature']}** ({r['PctContribution']:.0f}%, value: {r['Value']:.2f})"
                                        for _, r in top_pos.iterrows()
                                    )
                                    st.markdown(f"🔴 The strongest factors **pushing toward** "
                                                f"**{pred_class}** were: {reasons}.")
                                if not top_neg.empty:
                                    reasons2 = ", ".join(
                                        f"**{r['Feature']}** ({r['PctContribution']:.0f}%, value: {r['Value']:.2f})"
                                        for _, r in top_neg.iterrows()
                                    )
                                    st.markdown(f"🟢 The factors **working against** this prediction "
                                                f"were: {reasons2}.")

                                st.markdown("#### 📋 Full Feature Impact Table")
                                show_shap = shap_df[['Feature','Value','SHAP_Value','PctContribution']].copy()
                                show_shap['SHAP_Value']      = show_shap['SHAP_Value'].round(4)
                                show_shap['Value']           = show_shap['Value'].round(3)
                                show_shap['PctContribution'] = show_shap['PctContribution'].round(1)
                                show_shap.columns = ['Feature','Recorded Value','SHAP Impact','Contribution (%)']
                                st.dataframe(show_shap, use_container_width=True, hide_index=True)

                            except Exception as e:
                                st.error(f"SHAP explanation failed: {e}")
                                st.caption(
                                    "This usually means District_Model_Features.pkl doesn't match what the "
                                    "model expects, or the model doesn't support predict_proba."
                                )
        st.markdown('</div>', unsafe_allow_html=True)


    # ── FOOTER ────────────────────────────────────────────────────────────────
    st.markdown("""
    <div style='text-align:center;margin-top:20px;padding:14px 0;
                border-top:1px solid #CBD5E0;
                color:#4A5568;font-size:0.69rem;'>
        Telangana AI-Based District Drought Prediction &amp; Risk Assessment System &nbsp;·&nbsp;
        Department of Agriculture, Government of Telangana &nbsp;·&nbsp; District Model v5.0<br>
        <i>Official use only. Predictions must be validated by field officers before policy action.</i>
    </div>""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
#  ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────
def main():
    if "logged_in" not in st.session_state:
        st.session_state["logged_in"]=False
    if not st.session_state["logged_in"]:
        show_login()
    else:
        show_dashboard()

if __name__=="__main__":
    main()